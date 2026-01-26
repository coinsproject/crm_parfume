from datetime import datetime, date, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:
    # Для Python < 3.9 используем pytz (нужно установить: pip install pytz)
    try:
        import pytz
        ZoneInfo = None
    except ImportError:
        pytz = None
        ZoneInfo = None
import asyncio
from decimal import Decimal, InvalidOperation
import re
import time
from typing import List, Dict, Any, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Request, UploadFile, File, HTTPException, Form, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, OperationalError
import sqlalchemy as sa
import sqlite3
import httpx

from app.db import get_db
from app.models import User, PriceProduct, PriceHistory, PriceUpload, OrderItem, Client, Brand, BrandAlias, Partner
from app.services.auth_service import require_permission, user_has_permission, require_roles
from app.services.catalog_sync_service import sync_catalog_from_price
from app.services.partner_pricing_service import (
    get_partner_pricing_policy,
    get_total_markup_percent,
    calc_client_price,
    calc_partner_price,
)
from app.services.price_normalization_service import normalize_price_row, NormalizedResult
from app.services.catalog_upsert_service import upsert_catalog_from_price
from app.services.key_normalization import normalize_key
from app.services.catalog_background_service import create_catalog_items_from_price_batch
from app.logging_config import price_logger
from app.config_filters import PRICE_FILTERS, SECTIONS, PARFUM_KEYWORDS, PARFUM_FILTERS, COSMETICS_EXCLUSION_KEYWORDS
import json

router = APIRouter(prefix="/price", tags=["price"])
templates = Jinja2Templates(directory="app/templates")

# Добавляем фильтр from_json для шаблонов
def from_json_filter(value):
    """Фильтр Jinja2 для парсинга JSON"""
    if not value:
        return {}
    try:
        return json.loads(value) if isinstance(value, str) else value
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}

templates.env.filters["from_json"] = from_json_filter


@router.get("/upload/status/{upload_id}", response_class=JSONResponse)
async def get_upload_status(
    upload_id: int,
    current_user: User = Depends(require_permission("price.upload")),
    db: Session = Depends(get_db),
):
    """Получить статус загрузки прайса"""
    upload = db.query(PriceUpload).filter(PriceUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Загрузка не найдена")
    
    return {
        "id": upload.id,
        "status": upload.status,
        "processed_rows": upload.processed_rows or 0,
        "total_rows": upload.total_rows or 0,
        "progress_percent": float(upload.progress_percent or 0),
        "cancelled": upload.cancelled or False,
        "added_count": upload.added_count or 0,
        "updated_price_count": upload.updated_price_count or 0,
        "unchanged_count": upload.unchanged_count or 0,
    }


@router.get("/upload/latest", response_class=JSONResponse)
async def get_latest_upload(
    current_user: User = Depends(require_permission("price.upload")),
    db: Session = Depends(get_db),
):
    """Получить последнюю загрузку прайса"""
    upload = db.query(PriceUpload).order_by(PriceUpload.uploaded_at.desc()).first()
    if not upload:
        return {"id": None, "status": "none"}
    
    return {
        "id": upload.id,
        "status": upload.status,
        "processed_rows": upload.processed_rows or 0,
        "total_rows": upload.total_rows or 0,
        "progress_percent": float(upload.progress_percent or 0),
        "cancelled": upload.cancelled or False,
    }


@router.post("/upload/cancel/{upload_id}", response_class=JSONResponse)
async def cancel_upload(
    upload_id: int,
    current_user: User = Depends(require_permission("price.upload")),
    db: Session = Depends(get_db),
):
    """Отменить загрузку прайса"""
    upload = db.query(PriceUpload).filter(PriceUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Загрузка не найдена")
    
    if upload.status in ("done", "failed", "cancelled"):
        return {"message": "Загрузка уже завершена", "status": upload.status}
    
    upload.cancelled = True
    upload.status = "cancelled"
    db.commit()
    
    return {"message": "Загрузка отменена", "status": "cancelled"}


def _parse_raw_name(raw: str) -> Dict[str, Any]:
    """Parse brand/volume/gender/category/product_name from raw_name without mutating raw."""
    if not raw:
        return {
            "brand": "",
            "product_name": "",
            "category": "",
            "volume_value": None,
            "volume_unit": None,
            "gender": None,
        }

    # 2.1 нормализация
    raw_norm = " ".join(raw.strip().split())

    # 2.2 объем
    vol_match = re.search(r"(\d+(?:[.,]\d+)?)(\s*)(мл|ml|г|гр|g)", raw_norm, re.IGNORECASE)
    volume_value = None
    volume_unit = None
    if vol_match:
        volume_value = Decimal(vol_match.group(1).replace(",", "."))
        unit_raw = vol_match.group(3).lower()
        if unit_raw in ["мл", "ml"]:
            volume_unit = "мл"
        elif unit_raw in ["г", "гр", "g"]:
            volume_unit = "г"

    raw_low = raw_norm.lower()

    # 2.3 gender
    gender = None
    if "жен" in raw_low:
        gender = "F"
    elif "муж" in raw_low:
        gender = "M"
    elif "унис" in raw_low:
        gender = "U"

    # 2.5 category
    category = ""
    category_keywords = [
        ("шампунь против перхоти", "Уход за волосами"),
        ("парфюмированная вода", "Парфюм"),
        ("туалетная вода", "Парфюм"),
        ("духи", "Парфюм"),
        ("шампунь", "Уход за волосами"),
        ("маска для волос", "Уход за волосами"),
        ("спрей для волос", "Уход за волосами"),
        ("губная помада", "Декоративная косметика"),
        ("тональный крем", "Декоративная косметика"),
        ("пудра", "Декоративная косметика"),
        ("сыворотка для лица", "Уход за кожей"),
        ("сыворотка для глаз", "Уход за кожей"),
        ("крем для лица", "Уход за кожей"),
    ]
    for kw, cat in category_keywords:
        if kw in raw_low:
            category = cat
            break

    # 2.4 бренд
    brand = ""
    product_name = raw_norm

    def detect_brand(text: str) -> str:
        # Если иерархия с '>' - берём первый сегмент после удаления ведущих чисел/мусора
        if ">" in text:
            segments = [seg.strip() for seg in text.split(">") if seg.strip()]
            if segments:
                seg = re.sub(r"^[0-9]+\s*", "", segments[0])
                seg = re.sub(r"^[^A-Za-zА-Яа-яЁё]+", "", seg)
                first_token = seg.split()[0] if seg else ""
                return first_token.title()
        tokens = text.split()
        brand_tokens = []
        started = False
        for tok in tokens:
            if not started and tok.isdigit():
                continue
            started = True
            if re.search(r"[а-яё]", tok, re.IGNORECASE):
                break
            if tok.lower() in {"женский", "мужской", "унисекс"}:
                break
            brand_tokens.append(tok)
        if brand_tokens:
            return " ".join(brand_tokens)
        return tokens[0] if tokens else ""

    brand = detect_brand(raw_norm)

    # 2.6 product_name: убираем бренд в начале, гендерные слова и объем
    product_part = raw_norm
    if ">" in raw_norm:
        segments = [seg.strip() for seg in raw_norm.split(">") if seg.strip()]
        if segments:
            product_part = segments[-1]
    if brand and product_part.lower().startswith(brand.lower()):
        product_part = product_part[len(brand):].strip()
    # убираем объемное совпадение
    if vol_match:
        product_part = re.sub(r"[ ,]*\d+(?:[.,]\d+)?\s*(мл|ml|г|гр|g)$", "", product_part, flags=re.IGNORECASE)
    product_part = re.sub(r"[ \t,\\.]+$", "", product_part)
    # убираем половые маркеры
    for marker in ["женский", "мужской", "унисекс"]:
        product_part = re.sub(marker, "", product_part, flags=re.IGNORECASE)
    product_part = " ".join(product_part.split()).strip()
    if not product_part:
        product_part = raw_norm

    return {
        "brand": brand,
        "product_name": product_part,
        "category": category,
        "volume_value": volume_value,
        "volume_unit": volume_unit,
        "gender": gender,
    }


def _get_latest_price(price_product: PriceProduct) -> Decimal:
    if not price_product.price_history:
        return Decimal(0)
    last = sorted(price_product.price_history, key=lambda h: h.created_at or datetime.min)[-1]
    return Decimal((last.new_price_2 if last.new_price_2 is not None else last.price) or 0)


def _calc_price_fields(price_1: Decimal) -> (Decimal, Decimal):
    """Возвращает цену без округления и delta=0."""
    price_2 = price_1
    round_delta = Decimal("0")
    return price_2, round_delta


def _parse_decimal(value) -> Decimal:
    if value is None:
        return None
    try:
        num = Decimal(str(value).replace(" ", "").replace("\u00a0", ""))
        return num.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _ensure_price_indexes(db: Session):
    # ����� ���������, ���� �� ����쭮� ����� �� �� �����.
    db.execute(sa.text("CREATE INDEX IF NOT EXISTS idx_price_products_article ON price_products(external_article)"))
    db.execute(sa.text("CREATE INDEX IF NOT EXISTS idx_price_products_active ON price_products(is_active)"))
    db.execute(sa.text("CREATE INDEX IF NOT EXISTS idx_price_products_brand ON price_products(brand)"))
    db.execute(sa.text("CREATE INDEX IF NOT EXISTS idx_price_products_category ON price_products(category)"))
    db.execute(sa.text("CREATE INDEX IF NOT EXISTS idx_price_history_upload ON price_history(price_upload_id)"))
    db.execute(
        sa.text(
            "CREATE INDEX IF NOT EXISTS idx_price_history_pid_created "
            "ON price_history(price_product_id, created_at DESC)"
        )
    )


@router.get("/", response_class=HTMLResponse)
async def price_index(
    request: Request,
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    """Главная страница прайса - перенаправляет на поиск товаров"""
    # Всегда перенаправляем на страницу поиска товаров
    return RedirectResponse(url="/price/search", status_code=303)


@router.get("/upload_page", response_class=HTMLResponse)
async def price_upload_page(
    request: Request,
    current_user: User = Depends(require_roles(["ADMIN"])),
    ct: str = None,
    page: int = 1,
    db: Session = Depends(get_db),
):
    t0 = time.perf_counter()
    price_logger.info("[PRICE_UPLOAD_PAGE] start")

    t1 = time.perf_counter()
    uploads = (
        db.query(PriceUpload)
        .order_by(PriceUpload.uploaded_at.desc())
        .limit(10)
        .all()
    )
    price_logger.info("[PRICE_UPLOAD_PAGE] uploads loaded in %.3f sec", time.perf_counter() - t1)
    latest_upload = uploads[0] if uploads else None
    items_by_change: Dict[str, List[PriceHistory]] = {"NEW": [], "UP": [], "DOWN": [], "REMOVED": []}
    counts_by_change: Dict[str, int] = {"NEW": 0, "UP": 0, "DOWN": 0, "REMOVED": 0, "UNCHANGED": 0}
    pages_count = 1
    page = max(int(page or 1), 1)
    selected_change = ct if ct in items_by_change else "NEW"
    PAGE_SIZE = 50
    if latest_upload:
        latest_rows = db.query(
            PriceHistory.id,
            PriceHistory.price_product_id,
            PriceHistory.new_price_1,
            PriceHistory.old_price_1,
            PriceHistory.new_price_2,
            PriceHistory.old_price_2,
            PriceHistory.price,
        ).filter(PriceHistory.price_upload_id == latest_upload.id).all()

        prev_upload = (
            db.query(PriceUpload).filter(PriceUpload.id != latest_upload.id).order_by(PriceUpload.uploaded_at.desc()).first()
        )
        prev_map: Dict[int, Decimal] = {}
        if prev_upload:
            prev_rows = db.query(
                PriceHistory.price_product_id,
                PriceHistory.new_price_1,
            ).filter(PriceHistory.price_upload_id == prev_upload.id).all()
            for r in prev_rows:
                prev_map[r.price_product_id] = Decimal(str(r.new_price_1)) if r.new_price_1 is not None else None

        buckets: Dict[str, List[int]] = {k: [] for k in items_by_change.keys()}
        unchanged_ids: List[int] = []
        for r in latest_rows:
            new_price_1 = Decimal(str(r.new_price_1)) if r.new_price_1 is not None else None
            old_price_1_prev = prev_map.get(r.price_product_id)
            if old_price_1_prev is None:
                bucket = "NEW"
            elif new_price_1 is None:
                bucket = "REMOVED"
            else:
                if new_price_1 > (old_price_1_prev or Decimal(0)):
                    bucket = "UP"
                elif new_price_1 < (old_price_1_prev or Decimal(0)):
                    bucket = "DOWN"
                else:
                    bucket = "UNCHANGED"
            if bucket in buckets:
                buckets[bucket].append(r.id)
            elif bucket == "UNCHANGED":
                unchanged_ids.append(r.id)

        counts_by_change["NEW"] = len(buckets["NEW"])
        counts_by_change["UP"] = len(buckets["UP"])
        counts_by_change["DOWN"] = len(buckets["DOWN"])
        counts_by_change["REMOVED"] = len(buckets["REMOVED"])
        counts_by_change["UNCHANGED"] = len(unchanged_ids)

        def _changes(ct_val: str, offset: int = 0, limit: int = PAGE_SIZE):
            ids = buckets.get(ct_val, [])
            slice_ids = ids[offset : offset + limit]
            if not slice_ids:
                return []
            return (
                db.query(PriceHistory)
                .options(sa.orm.joinedload(PriceHistory.price_product))
                .filter(PriceHistory.id.in_(slice_ids))
                .order_by(PriceHistory.id.desc())
                .all()
            )

        t2 = time.perf_counter()
        total = counts_by_change.get(selected_change, 0)
        pages_count = (total + PAGE_SIZE - 1) // PAGE_SIZE if total else 1
        offset = (page - 1) * PAGE_SIZE
        items_by_change[selected_change] = _changes(selected_change, offset=offset, limit=PAGE_SIZE)
        price_logger.info("[PRICE_UPLOAD_PAGE] history chunks loaded in %.3f sec", time.perf_counter() - t2)

    t3 = time.perf_counter()
    resp = templates.TemplateResponse(
        "price_upload.html",
        {
            "request": request,
            "current_user": current_user,
        "uploads": uploads,
        "latest_upload": latest_upload,
        "items_by_change": items_by_change,
        "counts_by_change": counts_by_change,
        "selected_change": selected_change,
        "page": page,
        "pages_count": pages_count,
        "can_upload": True,
            "active_menu": "normalization",
    },
)
    price_logger.info("[PRICE_UPLOAD_PAGE] template rendered in %.3f sec", time.perf_counter() - t3)
    price_logger.info("[PRICE_UPLOAD_PAGE] total time %.3f sec", time.perf_counter() - t0)
    return resp


@router.post("/upload", response_class=RedirectResponse)
async def upload_price(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles(["ADMIN"])),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        price_logger.exception("openpyxl not installed: %s", e)
        raise HTTPException(status_code=500, detail="openpyxl не установлен, загрузка XLSX недоступна")

    upload = None
    try:
        # Проверяем целостность базы данных перед началом загрузки
        try:
            integrity_check = db.execute(sa.text("PRAGMA integrity_check")).scalar()
            if integrity_check != "ok":
                price_logger.error("[PRICE_UPLOAD] Database integrity check failed: %s", integrity_check)
                raise HTTPException(
                    status_code=500, 
                    detail=f"База данных повреждена. Проверьте целостность данных. Ошибка: {integrity_check[:200]}"
                )
        except Exception as integrity_error:
            # Если проверка не удалась, логируем, но продолжаем (может быть не SQLite)
            error_str = str(integrity_error)
            if "malformed" in error_str.lower() or "database disk image" in error_str.lower():
                price_logger.error("[PRICE_UPLOAD] Database corruption detected during integrity check: %s", error_str)
                raise HTTPException(
                    status_code=500,
                    detail="База данных повреждена. Выполните проверку целостности и восстановите из резервной копии."
                )
            price_logger.warning("[PRICE_UPLOAD] Could not check database integrity: %s", integrity_error)
        
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content))
        preferred_sheets = [
            "Позиции",
            "Флаконы",
        ]
        sheet_name = None
        ws = None
        for name in preferred_sheets:
            if name in wb.sheetnames:
                sheet_name = name
                ws = wb[name]
                break
        if ws is None:
            if not wb.sheetnames:
                raise HTTPException(status_code=400, detail="Файл не содержит листов")
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="Пустой файл")

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        def _idx(name: str, default: int):
            return headers.index(name) if name in headers else default

        idx_article = _idx("Артикул", 0)
        idx_name = _idx("Наименование", 1)
        idx_price = _idx("Цена", 2)

        upload = PriceUpload(
            filename=file.filename,
            source_date=date.today(),
            status="in_progress",
            created_by_user_id=current_user.id if current_user else None,
        )
        db.add(upload)
        db.flush()

        prev_in_pricelist_ids = {
            pid for (pid,) in db.query(PriceProduct.id).filter(PriceProduct.is_in_current_pricelist.is_(True)).all()
        }
        if not prev_in_pricelist_ids:
            prev_in_pricelist_ids = {
                pid for (pid,) in db.query(PriceProduct.id).filter(PriceProduct.is_active.is_(True)).all()
            }

        def _safe_parse(raw: str) -> Dict[str, Any]:
            try:
                return _parse_raw_name(raw or "")
            except Exception as e:
                price_logger.exception("Parse raw_name failed for '%s': %s", raw, e)
                return {
                    "brand": None,
                    "product_name": raw or "",
                    "category": None,
                    "volume_value": None,
                    "volume_unit": None,
                    "gender": None,
                }

        # Подсчитываем общее количество строк для прогресса
        total_file_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        upload.total_rows = total_file_rows
        db.flush()
        
        # ОТКЛЮЧАЕМ FTS ТРИГГЕРЫ ОДИН РАЗ В НАЧАЛЕ для ускорения загрузки
        fts_triggers_disabled = False
        try:
            db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_insert"))
            db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_update"))
            db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_delete"))
            fts_triggers_disabled = True
            price_logger.info("[PRICE_UPLOAD] FTS triggers disabled for bulk upload")
        except Exception as trigger_error:
            price_logger.warning("[PRICE_UPLOAD] Could not disable FTS triggers: %s", trigger_error)

        total_rows = 0
        added_count = updated_price_count = unchanged_count = 0
        up_count = down_count = 0
        marked_out_of_stock_count = 0
        seen_product_ids: List[int] = []
        seen_articles_in_file: Dict[str, int] = {}  # Отслеживаем артикулы в текущем файле для обработки дубликатов
        duplicate_articles_count = 0
        
        # Батчинг для оптимизации: коммитим каждые BATCH_SIZE строк
        # 100 строк - компромисс между скоростью и частотой обновления прогресса
        BATCH_SIZE = 100
        batch_count = 0
        rows_seen = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Обновляем прогресс и коммитим батч, если достигнут лимит
            if batch_count >= BATCH_SIZE:
                db.refresh(upload)
                if upload.cancelled:
                    upload.status = "cancelled"
                    upload.processed_rows = rows_seen
                    upload.progress_percent = (rows_seen / total_file_rows * 100) if total_file_rows > 0 else 0
                    db.commit()
                    raise HTTPException(status_code=200, detail="Загрузка отменена пользователем")
                upload.processed_rows = rows_seen
                upload.progress_percent = (rows_seen / total_file_rows * 100) if total_file_rows > 0 else 0
                db.commit()
                await asyncio.sleep(0)
                batch_count = 0

            # Проверяем отмену загрузки только в батчах (не на каждой строке)
            if batch_count == 0:
                db.refresh(upload)
                if upload.cancelled:
                    upload.status = "cancelled"
                    upload.processed_rows = rows_seen
                    upload.progress_percent = (rows_seen / total_file_rows * 100) if total_file_rows > 0 else 0
                    # Восстанавливаем FTS триггеры при отмене (только если FTS таблица не повреждена)
                    if fts_triggers_disabled:
                        try:
                            # Проверяем доступность FTS таблицы
                            try:
                                db.execute(sa.text("SELECT COUNT(*) FROM price_products_fts5 LIMIT 1")).scalar()
                                fts_table_ok = True
                            except (OperationalError, sqlite3.DatabaseError):
                                fts_table_ok = False
                            
                            if fts_table_ok:
                                db.execute(sa.text("""
                                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_insert AFTER INSERT ON price_products BEGIN
                                        INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                        VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                    END;
                                """))
                                db.execute(sa.text("""
                                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_update AFTER UPDATE ON price_products BEGIN
                                        DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                        INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                        VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                    END;
                                """))
                                db.execute(sa.text("""
                                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_delete AFTER DELETE ON price_products BEGIN
                                        DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                    END;
                                """))
                        except Exception:
                            pass
                    db.commit()
                    price_logger.info("[PRICE_UPLOAD] Upload cancelled by user at row %d", total_rows)
                    raise HTTPException(status_code=200, detail="Загрузка отменена пользователем")
            # Учитываем просмотренную строку для прогресса
            rows_seen += 1
            batch_count += 1
            if not row or len(row) <= idx_article:
                continue
            external_article = row[idx_article]
            if external_article in (None, "", "None"):
                continue
            external_article = str(external_article).strip()
            if external_article.lower().startswith("артикул"):
                continue
            
            # Проверяем дубликаты артикулов в текущем файле
            if external_article in seen_articles_in_file:
                duplicate_articles_count += 1
                price_logger.warning("[PRICE_UPLOAD] duplicate article in file: %s (row %d, first seen at row %d)", 
                                    external_article, total_rows + 1, seen_articles_in_file[external_article])
                # Пропускаем дубликат - используем данные из первой встречи артикула
                continue
            raw_name = str(row[idx_name]).strip() if len(row) > idx_name and row[idx_name] is not None else ""
            if raw_name.strip().lower().startswith("наименован"):
                continue
            price_val = row[idx_price] if len(row) > idx_price else None
            price_1 = _parse_decimal(price_val)
            if price_1 is None:
                price_logger.warning("[PRICE_UPLOAD] skip row without valid price: article=%s raw=%s val=%s", external_article, raw_name, price_val)
                continue

            price_2, round_delta = _calc_price_fields(price_1)
            parsed = _safe_parse(raw_name)
            now = datetime.utcnow()

            # Нормализация прайса ОТКЛЮЧЕНА при загрузке для ускорения
            # Нормализация будет запускаться вручную через кнопку "Прогнать нормализацию"
            normalized = None

            product = db.query(PriceProduct).filter(PriceProduct.external_article == external_article).first()
            if product:
                if product.price_2 is None:
                    last_hist = (
                        db.query(PriceHistory)
                        .filter(PriceHistory.price_product_id == product.id)
                        .order_by(PriceHistory.created_at.desc())
                        .first()
                    )
                    if last_hist:
                        fallback_price = last_hist.new_price_2 if last_hist.new_price_2 is not None else last_hist.price
                        product.price_2 = fallback_price
                        product.price_1 = last_hist.new_price_1 if last_hist.new_price_1 is not None else product.price_1
                old_price_1 = Decimal(product.price_1) if product.price_1 is not None else None
                old_price_2 = Decimal(product.price_2) if product.price_2 is not None else None
                old_round_delta = Decimal(product.round_delta) if product.round_delta is not None else None
                # определяем изменения по price_2
                if old_price_2 is None or price_2 > old_price_2:
                    change_type = "UP"
                    updated_price_count += 1
                    price_changed = True
                elif price_2 < (old_price_2 or Decimal(0)):
                    change_type = "DOWN"
                    updated_price_count += 1
                    price_changed = True
                else:
                    change_type = "UNCHANGED"
                    unchanged_count += 1
                    price_changed = False

                product.raw_name = raw_name
                product.product_name = parsed.get("product_name") or raw_name
                if parsed.get("brand"):
                    product.brand = parsed["brand"]
                if parsed.get("category"):
                    product.category = parsed["category"]
                if parsed.get("volume_value") is not None:
                    product.volume_value = parsed["volume_value"]
                if parsed.get("volume_unit"):
                    product.volume_unit = parsed["volume_unit"]
                if parsed.get("gender"):
                    product.gender = parsed["gender"]
                
                # Поля нормализации не заполняются при загрузке (для ускорения)
                # Устанавливаем статус "pending" - нормализация будет запущена вручную
                product.ai_status = "pending"
                product.normalization_notes = None
                
                product.is_active = True
                product.is_in_stock = True
                product.is_in_current_pricelist = True
                product.price_1 = price_1
                product.price_2 = price_2
                product.round_delta = round_delta
                if price_changed:
                    product.last_price_change_at = now

                if change_type == "UP":
                    up_count += 1
                elif change_type == "DOWN":
                    down_count += 1

                # Обновляем продукт через прямой SQL (триггеры FTS уже отключены)
                try:
                    db.execute(
                        sa.text("""
                            UPDATE price_products 
                            SET raw_name=:raw_name, product_name=:product_name, brand=:brand, category=:category,
                                volume_value=:volume_value, volume_unit=:volume_unit, gender=:gender,
                                is_active=:is_active, is_in_stock=:is_in_stock, 
                                is_in_current_pricelist=:is_in_current_pricelist,
                                price_1=:price_1, price_2=:price_2, round_delta=:round_delta,
                                last_price_change_at=:last_price_change_at,
                                norm_brand=:norm_brand, brand_confidence=:brand_confidence,
                                model_name=:model_name, series=:series,
                                category_path_json=:category_path_json, attrs_json=:attrs_json,
                                ai_group_key=:ai_group_key, variant_key=:variant_key,
                                search_text=:search_text, normalization_notes=:normalization_notes,
                                ai_status=:ai_status, updated_at=:updated_at
                            WHERE id=:id
                        """),
                        {
                            "id": product.id,
                            "raw_name": raw_name,
                            "product_name": parsed.get("product_name") or raw_name,
                            "brand": parsed.get("brand"),
                            "category": parsed.get("category"),
                            "volume_value": float(parsed.get("volume_value")) if parsed.get("volume_value") else None,
                            "volume_unit": parsed.get("volume_unit"),
                            "gender": parsed.get("gender"),
                            "is_active": True,
                            "is_in_stock": True,
                            "is_in_current_pricelist": True,
                            "price_1": float(price_1),
                            "price_2": float(price_2),
                            "round_delta": float(round_delta),
                            "last_price_change_at": now,
                            "norm_brand": None,
                            "brand_confidence": None,
                            "model_name": None,
                            "series": None,
                            "category_path_json": None,
                            "attrs_json": None,
                            "ai_group_key": None,
                            "variant_key": None,
                            "search_text": None,
                            "normalization_notes": None,
                            "ai_status": "pending",
                            "updated_at": now,
                        }
                    )
                    db.flush()
                except (OperationalError, sqlite3.DatabaseError) as db_error:
                    error_str = str(db_error)
                    # Если ошибка связана с FTS таблицей - просто пропускаем этот продукт
                    if "vtable" in error_str.lower() or "fts" in error_str.lower() or "malformed" in error_str.lower():
                        price_logger.warning("[PRICE_UPLOAD] Database error, skipping product update (article=%s): %s", 
                                            external_article, error_str[:200])
                        try:
                            db.rollback()
                        except:
                            pass
                        # Пропускаем этот продукт, но продолжаем загрузку остальных
                        continue
                    else:
                        # Другая ошибка БД - пробрасываем дальше
                        raise
                except Exception as update_error:
                    # Любая другая ошибка - логируем и пробрасываем
                    price_logger.error("[PRICE_UPLOAD] Unexpected error during product update (article=%s): %s", 
                                     external_article, update_error)
                    raise
                
                # Добавляем историю изменений
                try:
                    db.add(
                        PriceHistory(
                            price_product_id=product.id,
                            price=price_2,
                            old_price_1=old_price_1,
                        new_price_1=price_1,
                        old_price_2=old_price_2,
                        new_price_2=price_2,
                        old_round_delta=old_round_delta,
                        new_round_delta=round_delta,
                        currency="RUB",
                        source_date=date.today(),
                        source_filename=file.filename,
                        change_type=change_type,
                        price_upload_id=upload.id,
                        changed_at=now,
                    )
                    )
                except Exception as hist_error:
                    price_logger.warning("[PRICE_UPLOAD] Failed to add price history (article=%s): %s", 
                                        external_article, hist_error)
                    # Продолжаем без истории для этого продукта
                
                # Обновление каталога отключено при загрузке (будет выполнено после нормализации)
                
                # Записываем артикул в список обработанных
                seen_articles_in_file[external_article] = product.id
            else:
                product = PriceProduct(
                    external_article=external_article,
                    raw_name=raw_name,
                    product_name=parsed.get("product_name") or raw_name,
                    brand=parsed.get("brand"),
                    category=parsed.get("category"),
                    volume_value=parsed.get("volume_value"),
                    volume_unit=parsed.get("volume_unit"),
                    gender=parsed.get("gender"),
                    is_active=True,
                    is_in_stock=True,
                    is_in_current_pricelist=True,
                    price_1=price_1,
                    price_2=price_2,
                    round_delta=round_delta,
                    last_price_change_at=now,
                )
                
                # Поля нормализации не заполняются при загрузке (для ускорения)
                # Устанавливаем статус "pending" - нормализация будет запущена вручную
                product.ai_status = "pending"
                product.normalization_notes = None
                
                # Временно отключаем триггеры FTS перед созданием нового продукта
                triggers_disabled = False
                try:
                    db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_insert"))
                    db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_update"))
                    db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_delete"))
                    triggers_disabled = True
                    price_logger.debug("[PRICE_UPLOAD] FTS triggers temporarily disabled for new product")
                except Exception as trigger_error:
                    price_logger.warning("[PRICE_UPLOAD] Could not disable FTS triggers: %s", trigger_error)
                
                try:
                    db.add(product)
                    db.flush()  # Получаем ID продукта
                    
                    # Восстанавливаем триггеры FTS после успешного создания
                    if triggers_disabled:
                        try:
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_insert AFTER INSERT ON price_products BEGIN
                                    INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                    VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                END;
                            """))
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_update AFTER UPDATE ON price_products BEGIN
                                    DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                    INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                    VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                END;
                            """))
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_delete AFTER DELETE ON price_products BEGIN
                                    DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                END;
                            """))
                        except Exception as restore_error:
                            price_logger.warning("[PRICE_UPLOAD] Could not restore FTS triggers: %s", restore_error)
                    
                    seen_articles_in_file[external_article] = product.id
                    added_count += 1
                    up_count += 1  # новое считаем как повышение относительно 0 для статистики трендов
                except (IntegrityError, OperationalError, sqlite3.DatabaseError) as ie:
                    # Восстанавливаем триггеры в случае ошибки
                    if triggers_disabled:
                        try:
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_insert AFTER INSERT ON price_products BEGIN
                                    INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                    VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                END;
                            """))
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_update AFTER UPDATE ON price_products BEGIN
                                    DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                    INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                                    VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                                END;
                            """))
                            db.execute(sa.text("""
                                CREATE TRIGGER IF NOT EXISTS price_products_fts5_delete AFTER DELETE ON price_products BEGIN
                                    DELETE FROM price_products_fts5 WHERE rowid = old.id;
                                END;
                            """))
                        except Exception as restore_error:
                            price_logger.warning("[PRICE_UPLOAD] Could not restore FTS triggers in error handler: %s", restore_error)
                    
                    error_str = str(ie)
                    # Проверяем на ошибки FTS таблицы
                    if "vtable" in error_str.lower() or ("fts" in error_str.lower() and "constructor" in error_str.lower()):
                        price_logger.warning("[PRICE_UPLOAD] FTS error during product creation (article=%s), trying direct INSERT: %s", 
                                            external_article, error_str[:200])
                        db.rollback()
                        # Пытаемся создать продукт через прямой SQL, минуя триггеры FTS
                        try:
                            result = db.execute(
                                sa.text("""
                                    INSERT INTO price_products 
                                    (external_article, raw_name, product_name, brand, category, volume_value, volume_unit, gender,
                                     is_active, is_in_stock, is_in_current_pricelist, price_1, price_2, round_delta, last_price_change_at,
                                     norm_brand, brand_confidence, model_name, series, category_path_json, attrs_json, 
                                     ai_group_key, variant_key, search_text, normalization_notes, ai_status, created_at, updated_at)
                                    VALUES 
                                    (:external_article, :raw_name, :product_name, :brand, :category, :volume_value, :volume_unit, :gender,
                                     :is_active, :is_in_stock, :is_in_current_pricelist, :price_1, :price_2, :round_delta, :last_price_change_at,
                                     :norm_brand, :brand_confidence, :model_name, :series, :category_path_json, :attrs_json,
                                     :ai_group_key, :variant_key, :search_text, :normalization_notes, :ai_status, :created_at, :updated_at)
                                """),
                                {
                                    "external_article": external_article,
                                    "raw_name": raw_name,
                                    "product_name": parsed.get("product_name") or raw_name,
                                    "brand": parsed.get("brand"),
                                    "category": parsed.get("category"),
                                    "volume_value": float(parsed.get("volume_value")) if parsed.get("volume_value") else None,
                                    "volume_unit": parsed.get("volume_unit"),
                                    "gender": parsed.get("gender"),
                                    "is_active": True,
                                    "is_in_stock": True,
                                    "is_in_current_pricelist": True,
                                    "price_1": float(price_1),
                                    "price_2": float(price_2),
                                    "round_delta": float(round_delta),
                                    "last_price_change_at": now,
                                    "norm_brand": None,
                                    "brand_confidence": None,
                                    "model_name": None,
                                    "series": None,
                                    "category_path_json": None,
                                    "attrs_json": None,
                                    "ai_group_key": None,
                                    "variant_key": None,
                                    "search_text": None,
                                    "normalization_notes": None,
                                    "ai_status": "pending",
                                    "created_at": now,
                                    "updated_at": now,
                                }
                            )
                            product_id = result.lastrowid
                            product.id = product_id
                            seen_articles_in_file[external_article] = product_id
                            added_count += 1
                            up_count += 1
                            price_logger.info("[PRICE_UPLOAD] Product created via direct SQL (article=%s, id=%s)", external_article, product_id)
                        except Exception as direct_insert_error:
                            price_logger.error("[PRICE_UPLOAD] Failed to create product via direct SQL (article=%s): %s", 
                                             external_article, direct_insert_error)
                            # Пропускаем этот продукт
                            continue
                    # Проверяем на повреждение базы данных
                    elif "malformed" in error_str.lower() or "database disk image" in error_str.lower():
                        # Критическая ошибка - база данных повреждена
                        price_logger.error("[PRICE_UPLOAD] Database corruption detected: %s", error_str)
                        db.rollback()
                        raise HTTPException(
                            status_code=500,
                            detail="База данных повреждена. Выполните проверку целостности и восстановите из резервной копии."
                        )
                    # Если артикул уже существует (например, добавлен в параллельной транзакции)
                    else:
                        db.rollback()
                        db.flush()  # Сбрасываем состояние сессии
                        price_logger.warning("[PRICE_UPLOAD] article already exists, trying to update: %s", external_article)
                    # Пытаемся найти и обновить существующий продукт
                    existing_product = db.query(PriceProduct).filter(PriceProduct.external_article == external_article).first()
                    if existing_product:
                        # Обновляем существующий продукт вместо создания нового
                        product = existing_product
                        seen_articles_in_file[external_article] = product.id
                        # Переходим к логике обновления существующего продукта
                        if product.price_2 is None:
                            last_hist = (
                                db.query(PriceHistory)
                                .filter(PriceHistory.price_product_id == product.id)
                                .order_by(PriceHistory.created_at.desc())
                                .first()
                            )
                            if last_hist:
                                fallback_price = last_hist.new_price_2 if last_hist.new_price_2 is not None else last_hist.price
                                product.price_2 = fallback_price
                                product.price_1 = last_hist.new_price_1 if last_hist.new_price_1 is not None else product.price_1
                        old_price_1 = Decimal(product.price_1) if product.price_1 is not None else None
                        old_price_2 = Decimal(product.price_2) if product.price_2 is not None else None
                        old_round_delta = Decimal(product.round_delta) if product.round_delta is not None else None
                        # определяем изменения по price_2
                        if old_price_2 is None or price_2 > old_price_2:
                            change_type = "UP"
                            updated_price_count += 1
                            price_changed = True
                        elif price_2 < (old_price_2 or Decimal(0)):
                            change_type = "DOWN"
                            updated_price_count += 1
                            price_changed = True
                        else:
                            change_type = "UNCHANGED"
                            unchanged_count += 1
                            price_changed = False

                        product.raw_name = raw_name
                        product.product_name = parsed.get("product_name") or raw_name
                        if parsed.get("brand"):
                            product.brand = parsed["brand"]
                        if parsed.get("category"):
                            product.category = parsed["category"]
                        if parsed.get("volume_value") is not None:
                            product.volume_value = parsed["volume_value"]
                        if parsed.get("volume_unit"):
                            product.volume_unit = parsed["volume_unit"]
                        if parsed.get("gender"):
                            product.gender = parsed["gender"]
                        
                        # Поля нормализации не заполняются при загрузке (для ускорения)
                        # Устанавливаем статус "pending" - нормализация будет запущена вручную
                        product.ai_status = "pending"
                        product.normalization_notes = None
                        
                        product.is_active = True
                        product.is_in_stock = True
                        product.is_in_current_pricelist = True
                        product.price_1 = price_1
                        product.price_2 = price_2
                        product.round_delta = round_delta
                        if price_changed:
                            product.last_price_change_at = now

                        if change_type == "UP":
                            up_count += 1
                        elif change_type == "DOWN":
                            down_count += 1

                        # Пытаемся обновить продукт с обработкой ошибок FTS
                        try:
                            db.add(product)
                            db.flush()
                        except (OperationalError, sqlite3.DatabaseError) as fts_error:
                            error_str = str(fts_error)
                            if "vtable" in error_str.lower() or "fts" in error_str.lower() or "malformed" in error_str.lower():
                                price_logger.warning("[PRICE_UPLOAD] FTS error during product update in IntegrityError handler (article=%s): %s", 
                                                    external_article, error_str[:200])
                                db.rollback()
                                # Используем прямой UPDATE
                                try:
                                    db.execute(
                                        sa.text("""
                                            UPDATE price_products 
                                            SET is_active=:is_active, is_in_stock=:is_in_stock, 
                                                is_in_current_pricelist=:is_in_current_pricelist,
                                                price_1=:price_1, price_2=:price_2, round_delta=:round_delta,
                                                last_price_change_at=:last_price_change_at
                                            WHERE id=:id
                                        """),
                                        {
                                            "id": product.id,
                                            "is_active": True,
                                            "is_in_stock": True,
                                            "is_in_current_pricelist": True,
                                            "price_1": float(price_1),
                                            "price_2": float(price_2),
                                            "round_delta": float(round_delta),
                                            "last_price_change_at": now,
                                        }
                                    )
                                    db.flush()
                                except Exception as direct_update_error:
                                    price_logger.error("[PRICE_UPLOAD] Failed to update product via direct SQL in handler (article=%s): %s", 
                                                     external_article, direct_update_error)
                                    continue
                            else:
                                raise
                        
                        try:
                            db.add(
                                PriceHistory(
                                    price_product_id=product.id,
                                    price=price_2,
                                    old_price_1=old_price_1,
                                    new_price_1=price_1,
                                    old_price_2=old_price_2,
                                    new_price_2=price_2,
                                    old_round_delta=old_round_delta,
                                    new_round_delta=round_delta,
                                    currency="RUB",
                                    source_date=date.today(),
                                    source_filename=file.filename,
                                    change_type=change_type,
                                    price_upload_id=upload.id,
                                    changed_at=now,
                                )
                            )
                        except Exception as hist_error:
                            price_logger.warning("[PRICE_UPLOAD] Failed to add price history in handler (article=%s): %s", 
                                                external_article, hist_error)
                        
                        # Обновление каталога отключено при загрузке (будет выполнено после нормализации)
                        
                        seen_product_ids.append(product.id)
                        total_rows += 1
                        continue  # Пропускаем создание нового продукта, так как обновили существующий
                    else:
                        # Не удалось найти продукт, пропускаем строку
                        price_logger.error("[PRICE_UPLOAD] integrity error and product not found: %s", external_article)
                        continue
                
                # Обновление каталога отключено при загрузке (будет выполнено после нормализации)

                db.add(
                    PriceHistory(
                        price_product_id=product.id,
                        price=price_2,
                        old_price_1=None,
                        new_price_1=price_1,
                        old_price_2=None,
                        new_price_2=price_2,
                        old_round_delta=None,
                        new_round_delta=round_delta,
                        currency="RUB",
                        source_date=date.today(),
                        source_filename=file.filename,
                        change_type="NEW",
                        price_upload_id=upload.id,
                        changed_at=now,
                    )
                )

            # Записываем артикул в список обработанных (если еще не записан)
            if external_article not in seen_articles_in_file:
                seen_articles_in_file[external_article] = product.id
            seen_product_ids.append(product.id)
            total_rows += 1
            
        # REMOVED для отсутствующих
        removed_ids = set(prev_in_pricelist_ids) - set(seen_product_ids)
        if removed_ids:
            for prod in db.query(PriceProduct).filter(PriceProduct.id.in_(list(removed_ids))):
                old_price_1 = Decimal(prod.price_1) if prod.price_1 is not None else None
                old_price_2 = Decimal(prod.price_2) if prod.price_2 is not None else None
                old_round_delta = Decimal(prod.round_delta) if prod.round_delta is not None else None
                prod.is_active = False
                prod.is_in_stock = False
                prod.is_in_current_pricelist = False
                db.add(prod)
                db.add(
                    PriceHistory(
                        price_product_id=prod.id,
                        price=None,
                        old_price_1=old_price_1,
                        new_price_1=None,
                        old_price_2=old_price_2,
                        new_price_2=None,
                        old_round_delta=old_round_delta,
                        new_round_delta=None,
                        currency="RUB",
                        source_date=date.today(),
                        source_filename=file.filename,
                        change_type="REMOVED",
                        price_upload_id=upload.id,
                        changed_at=datetime.utcnow(),
                    )
                )
                marked_out_of_stock_count += 1
                down_count += 1

        # Финальное обновление прогресса
        upload.processed_rows = total_rows
        upload.progress_percent = 100.0
        upload.total_rows = total_rows
        upload.total_count = total_rows  # legacy
        upload.added_count = added_count
        upload.new_count = added_count  # legacy
        upload.updated_price_count = updated_price_count
        upload.up_count = up_count
        upload.down_count = down_count
        upload.marked_out_of_stock_count = marked_out_of_stock_count
        upload.removed_count = marked_out_of_stock_count  # legacy
        upload.unchanged_count = unchanged_count
        upload.status = "done"

        if duplicate_articles_count > 0:
            price_logger.warning("[PRICE_UPLOAD] found %d duplicate articles in file", duplicate_articles_count)

        # Синхронизация каталога отключена - будет выполняться в фоне после нормализации
        
        # ВОССТАНАВЛИВАЕМ FTS ТРИГГЕРЫ после завершения загрузки
        # Но только если FTS таблица не повреждена
        if fts_triggers_disabled:
            try:
                # Проверяем, что FTS таблица доступна, перед восстановлением триггеров
                try:
                    db.execute(sa.text("SELECT COUNT(*) FROM price_products_fts5 LIMIT 1")).scalar()
                    fts_table_ok = True
                except (OperationalError, sqlite3.DatabaseError) as fts_check_error:
                    error_str = str(fts_check_error)
                    if "vtable" in error_str.lower() or "fts" in error_str.lower() or "malformed" in error_str.lower():
                        price_logger.warning("[PRICE_UPLOAD] FTS table is corrupted, skipping trigger restoration: %s", error_str[:200])
                        fts_table_ok = False
                    else:
                        raise
                
                if fts_table_ok:
                    db.execute(sa.text("""
                        CREATE TRIGGER IF NOT EXISTS price_products_fts5_insert AFTER INSERT ON price_products BEGIN
                            INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                            VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                        END;
                    """))
                    db.execute(sa.text("""
                        CREATE TRIGGER IF NOT EXISTS price_products_fts5_update AFTER UPDATE ON price_products BEGIN
                            DELETE FROM price_products_fts5 WHERE rowid = old.id;
                            INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                            VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                        END;
                    """))
                    db.execute(sa.text("""
                        CREATE TRIGGER IF NOT EXISTS price_products_fts5_delete AFTER DELETE ON price_products BEGIN
                            DELETE FROM price_products_fts5 WHERE rowid = old.id;
                        END;
                    """))
                    price_logger.info("[PRICE_UPLOAD] FTS triggers restored after bulk upload")
                else:
                    price_logger.warning("[PRICE_UPLOAD] FTS triggers NOT restored due to corrupted FTS table. Search functionality may be limited.")
            except Exception as restore_error:
                price_logger.warning("[PRICE_UPLOAD] Could not restore FTS triggers: %s", restore_error)
        
        # Коммитим транзакцию с обработкой ошибок целостности
        try:
            db.commit()
        except IntegrityError as commit_error:
            error_str = str(commit_error)
            # Если ошибка связана с каталогом, откатываем и пробуем коммит без каталога
            if "catalog" in error_str.lower() or "price_product_id" in error_str.lower():
                price_logger.warning("[PRICE_UPLOAD] IntegrityError during commit (catalog related), rolling back catalog changes: %s", error_str[:200])
                db.rollback()
                # Пробуем коммит без синхронизации каталога
                try:
                    db.commit()
                    price_logger.info("[PRICE_UPLOAD] Committed without catalog sync due to IntegrityError")
                except Exception as retry_error:
                    price_logger.error("[PRICE_UPLOAD] Failed to commit after rollback: %s", retry_error)
                    raise
            else:
                # Другая ошибка целостности - пробрасываем
                raise
        
        price_logger.info(
            "Price upload success file=%s total=%s added=%s updated=%s unchanged=%s out_of_stock=%s",
            file.filename,
            upload.total_rows,
            upload.added_count,
            upload.updated_price_count,
            upload.unchanged_count,
            upload.marked_out_of_stock_count,
        )
        
        # Запускаем фоновую задачу для создания карточек каталога из нормализованных товаров
        # Это будет выполняться после завершения загрузки, не блокируя ответ
        background_tasks.add_task(create_catalog_items_from_price_batch, upload.id, batch_size=200)
        price_logger.info(f"[PRICE_UPLOAD] Upload {upload.id} completed. Background catalog creation started.")
        
        return RedirectResponse(url="/price/upload_page", status_code=303)
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        error_type = type(e).__name__
        price_logger.exception("Price upload failed for file=%s: %s (type=%s)", file.filename if file else "unknown", error_msg, error_type)
        
        # Определяем более конкретное сообщение об ошибке
        detail_msg = "Ошибка при загрузке прайса, попробуйте ещё раз"
        
        # Ошибки повреждения базы данных SQLite
        if "database disk image is malformed" in error_msg.lower() or "malformed" in error_msg.lower():
            detail_msg = "База данных повреждена. Выполните проверку целостности базы данных (PRAGMA integrity_check) и восстановите из резервной копии при необходимости."
        # Ошибки базы данных
        elif "database" in error_msg.lower() or "sql" in error_msg.lower() or "constraint" in error_msg.lower():
            detail_msg = "Ошибка базы данных при загрузке прайса. Проверьте целостность данных."
        # Ошибки транзакций
        elif "pendingrollback" in error_msg.lower() or "rollback" in error_msg.lower():
            detail_msg = "Ошибка транзакции базы данных. Попробуйте ещё раз или проверьте целостность базы данных."
        # Ошибки файла
        elif "file" in error_msg.lower() or "workbook" in error_msg.lower() or "sheet" in error_msg.lower():
            detail_msg = f"Ошибка чтения файла: {error_msg[:100]}"
        # Ошибки памяти
        elif "memory" in error_msg.lower() or "out of memory" in error_msg.lower():
            detail_msg = "Недостаточно памяти для обработки файла. Попробуйте загрузить файл меньшего размера."
        # Ошибки нормализации
        elif "normalization" in error_msg.lower() or "normalize" in error_msg.lower():
            detail_msg = f"Ошибка нормализации данных: {error_msg[:100]}"
        # Ошибки валидации данных
        elif "validation" in error_msg.lower() or "invalid" in error_msg.lower() or "value" in error_msg.lower():
            detail_msg = f"Ошибка валидации данных: {error_msg[:100]}"
        # Для отладки: показываем тип ошибки
        else:
            # В режиме разработки можно показывать больше информации
            detail_msg = f"Ошибка при загрузке прайса ({error_type}): {error_msg[:200]}"
        
        if upload:
            try:
                db.rollback()
                # фиксируем неуспешную попытку, чтобы было видно в истории
                failed_upload = PriceUpload(
                    filename=file.filename if file else None,
                    source_date=date.today(),
                    status="failed",
                    created_by_user_id=current_user.id if current_user else None,
                )
                db.add(failed_upload)
                db.commit()
            except Exception as rollback_error:
                price_logger.exception("Failed to save failed upload record: %s", rollback_error)
                db.rollback()
        else:
            try:
                db.rollback()
            except Exception as rollback_error:
                price_logger.exception("Failed to rollback transaction: %s", rollback_error)
        
        raise HTTPException(status_code=500, detail=detail_msg)


@router.post("/upload/{upload_id}/delete", response_class=RedirectResponse)
async def delete_price_upload(
    upload_id: int,
    current_user: User = Depends(require_permission("price.upload")),
    db: Session = Depends(get_db),
):
    upload = db.query(PriceUpload).filter(PriceUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Загрузка не найдена")

    # запомним затронутые товары
    affected_product_ids = [
        pid for (pid,) in db.query(PriceHistory.price_product_id).filter(PriceHistory.price_upload_id == upload.id).distinct()
    ]
    # удаляем связанные записи истории
    db.query(PriceHistory).filter(PriceHistory.price_upload_id == upload.id).delete()
    db.delete(upload)
    db.flush()

    if affected_product_ids:
        chunk_size = 800
        for i in range(0, len(affected_product_ids), chunk_size):
            chunk = affected_product_ids[i : i + chunk_size]
            sub = (
                db.query(
                    PriceHistory.price_product_id.label("pid"),
                    sa.func.max(PriceHistory.created_at).label("max_created"),
                )
                .filter(PriceHistory.price_product_id.in_(chunk))
                .group_by(PriceHistory.price_product_id)
                .subquery()
            )
            latest_rows = (
                db.query(PriceHistory)
                .join(sub, (PriceHistory.price_product_id == sub.c.pid) & (PriceHistory.created_at == sub.c.max_created))
                .all()
            )
            latest_map = {r.price_product_id: r for r in latest_rows}

            # Отключаем FTS триггеры один раз перед обновлением всех продуктов в чанке (FTS таблица может быть повреждена)
            try:
                db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_update"))
                db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_insert"))
                db.execute(sa.text("DROP TRIGGER IF EXISTS price_products_fts5_delete"))
            except Exception:
                pass  # Игнорируем ошибки при отключении триггеров

            # обновим только затронутые товары
            for pid in chunk:
                hist = latest_map.get(pid)
                
                try:
                    if hist:
                        ct = hist.change_type
                        is_active = bool(ct) and ct != "REMOVED"
                        db.query(PriceProduct).filter(PriceProduct.id == pid).update(
                        {
                            "is_active": is_active,
                            "is_in_stock": is_active,
                            "is_in_current_pricelist": is_active,
                            "price_1": hist.new_price_1,
                            "price_2": hist.new_price_2 or hist.price,
                            "round_delta": hist.new_round_delta,
                        }
                        )
                    else:
                        db.query(PriceProduct).filter(PriceProduct.id == pid).update(
                            {"is_active": False, "is_in_stock": False, "is_in_current_pricelist": False}
                        )
                except (OperationalError, sqlite3.DatabaseError) as db_error:
                    error_str = str(db_error)
                    if "vtable" in error_str.lower() or "fts" in error_str.lower():
                        # FTS таблица повреждена - используем прямой SQL
                        price_logger.warning("[PRICE_UPLOAD_DELETE] FTS error, using direct SQL for product_id=%s: %s", pid, error_str[:200])
                        try:
                            if hist:
                                ct = hist.change_type
                                is_active = bool(ct) and ct != "REMOVED"
                                db.execute(
                                    sa.text("""
                                        UPDATE price_products 
                                        SET is_active=:is_active, is_in_stock=:is_in_stock, 
                                            is_in_current_pricelist=:is_in_current_pricelist,
                                            price_1=:price_1, price_2=:price_2, round_delta=:round_delta,
                                            updated_at=:updated_at
                                        WHERE id=:id
                                    """),
                                    {
                                        "id": pid,
                                        "is_active": is_active,
                                        "is_in_stock": is_active,
                                        "is_in_current_pricelist": is_active,
                                        "price_1": hist.new_price_1,
                                        "price_2": hist.new_price_2 or hist.price,
                                        "round_delta": hist.new_round_delta,
                                        "updated_at": datetime.utcnow(),
                                    }
                                )
                            else:
                                db.execute(
                                    sa.text("""
                                        UPDATE price_products 
                                        SET is_active=:is_active, is_in_stock=:is_in_stock, 
                                            is_in_current_pricelist=:is_in_current_pricelist,
                                            updated_at=:updated_at
                                        WHERE id=:id
                                    """),
                                    {
                                        "id": pid,
                                        "is_active": False,
                                        "is_in_stock": False,
                                        "is_in_current_pricelist": False,
                                        "updated_at": datetime.utcnow(),
                                    }
                                )
                        except Exception as direct_sql_error:
                            price_logger.error("[PRICE_UPLOAD_DELETE] Direct SQL also failed for product_id=%s: %s", pid, direct_sql_error)
                            # Пропускаем этот продукт
                    else:
                        raise
            
            # Восстанавливаем FTS триггеры после обработки всех продуктов в чанке
            try:
                db.execute(sa.text("""
                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_update AFTER UPDATE ON price_products BEGIN
                        DELETE FROM price_products_fts5 WHERE rowid = old.id;
                        INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                        VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                    END;
                """))
                db.execute(sa.text("""
                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_insert AFTER INSERT ON price_products BEGIN
                        INSERT INTO price_products_fts5(rowid, raw_name, norm_brand, model_name, series, search_text, external_article)
                        VALUES (new.id, new.raw_name, new.norm_brand, new.model_name, new.series, new.search_text, new.external_article);
                    END;
                """))
                db.execute(sa.text("""
                    CREATE TRIGGER IF NOT EXISTS price_products_fts5_delete AFTER DELETE ON price_products BEGIN
                        DELETE FROM price_products_fts5 WHERE rowid = old.id;
                    END;
                """))
            except Exception:
                pass  # Игнорируем ошибки при восстановлении триггеров

    db.commit()
    return RedirectResponse(url="/price/upload_page", status_code=303)


@router.get("/upload/{upload_id}/delete", response_class=RedirectResponse)
async def delete_price_upload_get(
    upload_id: int,
    current_user: User = Depends(require_permission("price.upload")),
    db: Session = Depends(get_db),
):
    # удобный путь для удаления через ссылку
    return await delete_price_upload(upload_id, current_user, db)


@router.get("/search", response_class=HTMLResponse)
async def price_search_page(
    request: Request,
    q: str = "",
    page: int = 1,
    page_size: int = 20,
    upload_id: int = None,
    client_id: str | None = None,
    partner_id: str | None = None,
    brand: str | None = None,
    gender: str | None = None,
    ptype: str | None = None,
    psub: str | None = None,
    section: str | None = None,
    pf: str | None = None,
    hide_decant: bool = False,
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    try:
        t0 = time.perf_counter()
        price_logger.info("[PRICE_SEARCH] start q='%s' page=%s size=%s upload=%s", q, page, page_size, upload_id)
        page_size = max(10, min(int(page_size or 20), 20))
        can_view_cost = user_has_permission(current_user, db, "prices.view_cost")
        can_view_client = user_has_permission(current_user, db, "prices.view_client") or can_view_cost
        can_create_orders = user_has_permission(current_user, db, "orders.create")
        can_upload = user_has_permission(current_user, db, "price.upload")
        is_admin = current_user.role and current_user.role.name == 'ADMIN'
        # Преобразуем hide_decant в bool
        price_logger.info("[PRICE_SEARCH] Received hide_decant parameter: '%s' (type: %s)", hide_decant, type(hide_decant).__name__)
        # Обрабатываем разные варианты: строка "1", "true", "on", или уже bool
        if isinstance(hide_decant, bool):
            hide_decant_bool = hide_decant
        elif isinstance(hide_decant, str):
            hide_decant_bool = hide_decant.lower() in ("1", "true", "on", "yes")
        elif hide_decant is None:
            hide_decant_bool = False
        else:
            hide_decant_bool = bool(hide_decant)
        price_logger.info("[PRICE_SEARCH] Converted hide_decant_bool: %s", hide_decant_bool)
        t1 = time.perf_counter()
        current_upload = None
        upload_date_moscow = None
        if upload_id is None:
            latest_upload = db.query(PriceUpload).order_by(PriceUpload.uploaded_at.desc()).first()
            if latest_upload:
                upload_id = latest_upload.id
                current_upload = latest_upload
        else:
            current_upload = db.query(PriceUpload).filter(PriceUpload.id == upload_id).first()
        
        # Преобразуем дату в московское время
        if current_upload and current_upload.uploaded_at:
            try:
                if ZoneInfo:
                    # Python 3.9+
                    moscow_tz = ZoneInfo('Europe/Moscow')
                    # Если дата не имеет timezone, считаем её UTC
                    if current_upload.uploaded_at.tzinfo is None:
                        utc_dt = current_upload.uploaded_at.replace(tzinfo=timezone.utc)
                    else:
                        utc_dt = current_upload.uploaded_at.astimezone(timezone.utc)
                    upload_date_moscow = utc_dt.astimezone(moscow_tz)
                elif pytz:
                    # Python < 3.9, используем pytz
                    moscow_tz = pytz.timezone('Europe/Moscow')
                    # Если дата не имеет timezone, считаем её UTC
                    if current_upload.uploaded_at.tzinfo is None:
                        utc_dt = pytz.utc.localize(current_upload.uploaded_at)
                    else:
                        utc_dt = current_upload.uploaded_at.astimezone(pytz.utc)
                    upload_date_moscow = utc_dt.astimezone(moscow_tz)
                else:
                    # Fallback: просто добавляем 3 часа (UTC+3 для Москвы)
                    upload_date_moscow = current_upload.uploaded_at
            except Exception as e:
                price_logger.warning(f"[PRICE_SEARCH] Error converting timezone: {e}")
                upload_date_moscow = current_upload.uploaded_at
        
        price_logger.info("[PRICE_SEARCH] upload selection %.3f sec", time.perf_counter() - t1)

        t2 = time.perf_counter()
        try:
            products, total = _search_products(db, q, page, page_size, upload_id, brand=brand, gender=gender, ptype=ptype, psub=psub, section=section, pf=pf, hide_decant=hide_decant_bool)
        except Exception as search_error:
            price_logger.exception("[PRICE_SEARCH] Error in _search_products: %s", search_error)
            # В случае ошибки возвращаем пустой результат
            products = []
            total = 0
        price_logger.info("[PRICE_SEARCH] query products %.3f sec (total=%s)", time.perf_counter() - t2, total)
        
        # Вычисляем количество страниц
        pages_count = max(1, (total + page_size - 1) // page_size) if total > 0 else 1
    
        # Получаем список уникальных брендов для выпадающего списка
        # ОПТИМИЗАЦИЯ: Используем более быстрый запрос с LIMIT для больших таблиц
        brands_query = db.query(PriceProduct.brand).filter(
            PriceProduct.is_active.is_(True),
            PriceProduct.is_in_stock.is_(True),
            PriceProduct.is_in_current_pricelist.is_(True),
            PriceProduct.brand.isnot(None),
            PriceProduct.brand != ""
        ).distinct().order_by(PriceProduct.brand)
        
        # Ограничиваем количество брендов для производительности
        # Если брендов больше 1000, берем только первые 1000
        brands_result = brands_query.limit(1000).all()
        brands_list = [b[0] for b in brands_result if b[0]]
        
        # Получаем список партнёров для фильтра (только для админа)
        partners_list = []
        selected_partner_id = None
        selected_partner = None
        if is_admin:
            partners_list = db.query(Partner).order_by(Partner.name).all()
            if partner_id not in (None, "", "None", "null"):
                try:
                    selected_partner_id = int(partner_id)
                    selected_partner = db.query(Partner).filter(Partner.id == selected_partner_id).first()
                except ValueError:
                    selected_partner_id = None
        
        t3 = time.perf_counter()
        
        # Безопасная обработка истории и цен
        try:
            product_ids = [p.id for p in products] if products else []
            last_history = _last_history_map(db, product_ids) if product_ids else {}
            base_price_map = _latest_price_map(db, product_ids) if product_ids else {}
            raw_price_1_map = _latest_price_map(db, product_ids, field="price_1") if product_ids else {}
            price_map = base_price_map if can_view_cost else {}
            price_1_map = raw_price_1_map if can_view_cost else {}
        except Exception as history_error:
            price_logger.exception("[PRICE_SEARCH] Error loading history/price maps: %s", history_error)
            last_history = {}
            base_price_map = {}
            price_map = {}
            price_1_map = {}
            raw_price_1_map = {}
        
        price_logger.info("[PRICE_SEARCH] history maps %.3f sec", time.perf_counter() - t3)
        pages_count = (total + page_size - 1) // page_size if total else 1

        client_id_int = None
        if client_id not in (None, "", "None", "null"):
            try:
                client_id_int = int(client_id)
            except ValueError:
                client_id_int = None

        client = db.query(Client).filter(Client.id == client_id_int).first() if client_id_int else None

        partner_pricing = None
        # Определяем, для какого партнёра показывать цены
        view_partner_id = None
        admin_selected_partner_pricing = None
        
        # Если админ выбрал партнёра в фильтре, рассчитываем цену партнёра с учётом надбавки
        if is_admin and selected_partner_id:
            view_partner_id = selected_partner_id
            policy = get_partner_pricing_policy(db, selected_partner_id)
            # Рассчитываем цену партнёра с учётом надбавки (price_1 * (1 + partner_price_markup_percent / 100))
            admin_selected_partner_pricing = {
                pid: calc_partner_price(price, policy.partner_price_markup_percent)
                for pid, price in raw_price_1_map.items()
            }
        elif getattr(current_user, "partner_id", None):
            # Если текущий пользователь - партнёр, рассчитываем цены с наценками
            view_partner_id = current_user.partner_id
        
        if view_partner_id and can_view_client and not (is_admin and selected_partner_id):
            policy = get_partner_pricing_policy(db, view_partner_id)
            admin_pct = policy.admin_markup_percent
            total_pct = get_total_markup_percent(db, view_partner_id, client_id=client.id) if client else get_total_markup_percent(db, view_partner_id)
            partner_base_price_map = {
                pid: calc_partner_price(price, policy.partner_price_markup_percent)
                for pid, price in raw_price_1_map.items()
            }
            admin_price_map = {pid: calc_client_price(price, admin_pct) for pid, price in partner_base_price_map.items()}
            total_price_map = {pid: calc_client_price(price, total_pct) for pid, price in partner_base_price_map.items()}
            partner_pricing = {
                "admin_markup_percent": admin_pct,
                "partner_price_markup_percent": policy.partner_price_markup_percent,
                "partner_markup_percent": (total_pct - admin_pct) if total_pct is not None else None,
                "total_markup_percent": total_pct,
                "partner_default_markup_percent": policy.partner_default_markup_percent,
                "max_partner_markup_percent": policy.max_partner_markup_percent,
                "base_price_map": partner_base_price_map,
                "admin_price_map": admin_price_map,
                "total_price_map": total_price_map,
            }

        t4 = time.perf_counter()
        resp = templates.TemplateResponse(
            "price_search.html",
            {
                "request": request,
                "current_user": current_user,
                "active_menu": "price",
                "q": q,
                "products": products,
                "page": page,
                "page_size": page_size,
                "pages_count": pages_count,
                "total": total,
                "can_view_cost": can_view_cost,
                "can_view_client": can_view_client,
                "can_create_orders": can_create_orders,
                "price_map": price_map,
                "price_1_map": price_1_map,
                "last_history": last_history,
                "upload_id": upload_id,
                "client": client,
                "client_id": client.id if client else None,
                "partner_pricing": partner_pricing,
                "base_price_map": base_price_map if can_view_client else {},
                "brands_list": brands_list,
                "selected_brand": brand or "",
                "selected_gender": gender or "",
                "can_upload": can_upload,
                "ptype": ptype or "",
                "psub": psub or "",
                "price_filters": PRICE_FILTERS,
                "section": section,  # Передаем как есть (None или пустую строку), не заменяем на "parfum"
                "pf": pf or "",
                "sections": SECTIONS,
                "view_partner_id": view_partner_id,
                "is_admin": is_admin,
                "partners_list": partners_list,
                "selected_partner_id": selected_partner_id,
                "selected_partner": selected_partner,
                "admin_selected_partner_pricing": admin_selected_partner_pricing,
                "current_upload": current_upload,
                "upload_date_moscow": upload_date_moscow,
                "hide_decant": hide_decant_bool,
            },
        )
        price_logger.info("[PRICE_SEARCH] render %.3f sec", time.perf_counter() - t4)
        price_logger.info("[PRICE_SEARCH] total %.3f sec", time.perf_counter() - t0)
        return resp
    except Exception as e:
        price_logger.exception("[PRICE_SEARCH] Unexpected error: %s", e)
        try:
            db.rollback()
        except:
            pass
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при выполнении поиска: {str(e)[:200]}"
        )


@router.get("/api/brands", response_class=JSONResponse)
async def get_brands_api(
    q: str = "",
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    """Получить список брендов с поиском"""
    query = db.query(PriceProduct.brand).filter(
        PriceProduct.is_active.is_(True),
        PriceProduct.is_in_stock.is_(True),
        PriceProduct.is_in_current_pricelist.is_(True),
        PriceProduct.brand.isnot(None),
        PriceProduct.brand != ""
    ).distinct()
    
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.filter(PriceProduct.brand.ilike(search_term))
    
    brands = [b[0] for b in query.order_by(PriceProduct.brand).limit(100).all() if b[0]]
    return {"brands": brands}


@router.get("/api/dictionaries", response_class=JSONResponse)
async def get_dictionaries_api(
    field: str = "",
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    """Получить уникальные значения справочников из БД
    
    Параметры:
    - field: название поля (gender, brand, category и т.д.)
    """
    if not field:
        # Если поле не указано, возвращаем все доступные справочники
        return {
            "available_fields": ["gender", "brand", "category"],
            "message": "Укажите параметр field для получения значений конкретного справочника"
        }
    
    field = field.lower().strip()
    result = {
        "field": field,
        "values": [],
        "count": 0
    }
    
    if field == "gender":
        # Получаем уникальные значения пола
        query = db.query(PriceProduct.gender).filter(
            PriceProduct.is_active.is_(True),
            PriceProduct.is_in_stock.is_(True),
            PriceProduct.is_in_current_pricelist.is_(True),
            PriceProduct.gender.isnot(None),
            PriceProduct.gender != ""
        ).distinct()
        values = [v[0] for v in query.order_by(PriceProduct.gender).all() if v[0]]
        result["values"] = values
        result["count"] = len(values)
        result["labels"] = {
            "F": "Женский",
            "M": "Мужской",
            "U": "Унисекс"
        }
        
    elif field == "brand":
        # Получаем уникальные значения брендов
        query = db.query(PriceProduct.brand).filter(
            PriceProduct.is_active.is_(True),
            PriceProduct.is_in_stock.is_(True),
            PriceProduct.is_in_current_pricelist.is_(True),
            PriceProduct.brand.isnot(None),
            PriceProduct.brand != ""
        ).distinct()
        values = [v[0] for v in query.order_by(PriceProduct.brand).all() if v[0]]
        result["values"] = values
        result["count"] = len(values)
        
    elif field == "category":
        # Получаем уникальные значения категорий
        query = db.query(PriceProduct.category).filter(
            PriceProduct.is_active.is_(True),
            PriceProduct.is_in_stock.is_(True),
            PriceProduct.is_in_current_pricelist.is_(True),
            PriceProduct.category.isnot(None),
            PriceProduct.category != ""
        ).distinct()
        values = [v[0] for v in query.order_by(PriceProduct.category).all() if v[0]]
        result["values"] = values
        result["count"] = len(values)
        
    else:
        return JSONResponse(
            status_code=400,
            content={"error": f"Неизвестное поле: {field}. Доступные поля: gender, brand, category"}
        )
    
    return result


@router.get("/api/search", response_class=JSONResponse)
async def price_search_api(
    q: str = "",
    page: int = 1,
    page_size: int = 20,
    upload_id: int = None,
    partner_id: str | None = None,
    brand: str | None = None,
    gender: str | None = None,
    ptype: str | None = None,
    psub: str | None = None,
    section: str | None = None,
    pf: str | None = None,
    hide_decant: str | None = None,
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    page_size = max(10, min(int(page_size or 20), 20))
    can_view_cost = user_has_permission(current_user, db, "prices.view_cost")
    can_view_client = user_has_permission(current_user, db, "prices.view_client") or can_view_cost
    is_admin = current_user.role and current_user.role.name == 'ADMIN'
    # Преобразуем hide_decant в bool
    if isinstance(hide_decant, bool):
        hide_decant_bool = hide_decant
    elif isinstance(hide_decant, str):
        hide_decant_bool = hide_decant.lower() in ("1", "true", "on", "yes")
    elif hide_decant is None:
        hide_decant_bool = False
    else:
        hide_decant_bool = bool(hide_decant)
    if upload_id is None:
        latest_upload = db.query(PriceUpload).order_by(PriceUpload.uploaded_at.desc()).first()
        if latest_upload:
            upload_id = latest_upload.id
    products, total = _search_products(db, q, page, page_size, upload_id, brand=brand, gender=gender, ptype=ptype, psub=psub, section=section, pf=pf, hide_decant=hide_decant_bool)
    last_history = _last_history_map(db, [p.id for p in products])
    base_price_map = _latest_price_map(db, [p.id for p in products]) if can_view_client else {}
    raw_price_1_map = _latest_price_map(db, [p.id for p in products], field="price_1") if can_view_client else {}
    price_map = base_price_map if can_view_cost else {}
    price_1_map = raw_price_1_map if can_view_cost else {}
    
    # Определяем, для какого партнёра показывать цены
    view_partner_id = None
    admin_selected_partner_pricing = None
    
    if is_admin and partner_id not in (None, "", "None", "null"):
        try:
            selected_partner_id_for_api = int(partner_id)
            view_partner_id = selected_partner_id_for_api
            # Рассчитываем цену партнёра с учётом надбавки
            from app.services.partner_pricing_service import get_partner_pricing_policy, calc_partner_price
            from decimal import Decimal
            policy = get_partner_pricing_policy(db, selected_partner_id_for_api)
            admin_selected_partner_pricing = {
                p.id: float(calc_partner_price(Decimal(str(raw_price_1_map.get(p.id, 0))), policy.partner_price_markup_percent))
                for p in products if p.id in raw_price_1_map
            }
        except ValueError:
            pass
    elif getattr(current_user, "partner_id", None):
        view_partner_id = current_user.partner_id
    
    # Получаем информацию о партнере для расчета цен
    partner_pricing = None
    if view_partner_id and can_view_client and not (is_admin and partner_id not in (None, "", "None", "null")):
        from app.services.partner_pricing_service import (
            get_partner_pricing_policy,
            get_total_markup_percent,
            calc_client_price,
            calc_partner_price,
        )
        policy = get_partner_pricing_policy(db, view_partner_id)
        admin_pct = policy.admin_markup_percent
        total_pct = get_total_markup_percent(db, view_partner_id)
        partner_base_price_map = {
            pid: calc_partner_price(price, policy.partner_price_markup_percent)
            for pid, price in raw_price_1_map.items()
        }
        admin_price_map = {pid: calc_client_price(price, admin_pct) for pid, price in partner_base_price_map.items()}
        total_price_map = {pid: calc_client_price(price, total_pct) for pid, price in partner_base_price_map.items()}
        partner_pricing = {
            "admin_markup_percent": admin_pct,
            "partner_price_markup_percent": policy.partner_price_markup_percent,
            "total_markup_percent": total_pct,
            "base_price_map": partner_base_price_map,
            "admin_price_map": admin_price_map,
            "total_price_map": total_price_map,
        }
        base_price_map = partner_base_price_map
    
    return {
        "products": [
            {
                "id": p.id,
                "external_article": p.external_article,
                "raw_name": p.raw_name,
                "brand": p.brand,
                "product_name": p.product_name,
                "category": p.category,
                "volume_value": float(p.volume_value or 0),
                "volume_unit": p.volume_unit,
                "change_type": last_history.get(p.id).change_type if last_history.get(p.id) else None,
                "price": float(price_map.get(p.id, 0)) if can_view_cost and p.id in price_map else None,
                "price_1": float(price_1_map.get(p.id, 0)) if can_view_cost and p.id in price_1_map else None,
                "base_price": float(base_price_map.get(p.id, 0)) if can_view_client and p.id in base_price_map else None,
                "admin_price": float(partner_pricing["admin_price_map"].get(p.id, 0)) if partner_pricing and "admin_price_map" in partner_pricing and p.id in partner_pricing["admin_price_map"] else None,
                "total_price": float(partner_pricing["total_price_map"].get(p.id, 0)) if partner_pricing and "total_price_map" in partner_pricing and p.id in partner_pricing["total_price_map"] else None,
                "norm_brand": getattr(p, 'norm_brand', None),
                "model_name": getattr(p, 'model_name', None),
            }
            for p in products
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages_count": max(1, (total + page_size - 1) // page_size) if total > 0 else 1,
        "partner_pricing": partner_pricing,
        "can_view_cost": can_view_cost,
        "can_view_client": can_view_client,
        "is_admin": is_admin,
        "view_partner_id": view_partner_id,
        "admin_selected_partner_pricing": admin_selected_partner_pricing,
    }


def _normalize_text(text: str) -> str:
    """Нормализация текста: lower + замена ё→е"""
    if not text:
        return ""
    return text.lower().replace("ё", "е").replace("Ё", "е")


def _build_ilike_any(column, keywords: List[str]):
    """Строит OR фильтр по ключевым словам для ILIKE"""
    if not keywords:
        return None
    conditions = []
    for keyword in keywords:
        conditions.append(column.ilike(f"%{keyword}%"))
    if len(conditions) == 1:
        return conditions[0]
    # OR между всеми условиями
    from sqlalchemy import or_
    return or_(*conditions)


def _parse_csv_param(param: str | None) -> List[str]:
    """Парсит CSV параметр в список значений"""
    if not param or not param.strip():
        return []
    return [v.strip() for v in param.split(",") if v.strip()]


def _apply_section_filter(query, section: str):
    """
    Применяет фильтр по разделу (section)
    
    Для парфюма:
    - Позиция считается парфюмом, если содержит хотя бы одно из PARFUM_KEYWORDS
    - Исключение: если содержит косметические слова (COSMETICS_EXCLUSION_KEYWORDS),
      то приоритет у Косметики, даже если есть парфюм-слово
    """
    if section == "parfum":
        from sqlalchemy import or_, and_, not_
        
        # Условие 1: содержит хотя бы одно парфюм-слово
        parfum_conditions = []
        for keyword in PARFUM_KEYWORDS:
            parfum_conditions.append(PriceProduct.raw_name.ilike(f"%{keyword}%"))
        
        if not parfum_conditions:
            return query
        
        parfum_match = or_(*parfum_conditions)
        
        # Условие 2: НЕ содержит косметические слова (исключения)
        cosmetics_exclusions = []
        for exclusion in COSMETICS_EXCLUSION_KEYWORDS:
            cosmetics_exclusions.append(PriceProduct.raw_name.ilike(f"%{exclusion}%"))
        
        if cosmetics_exclusions:
            # Позиция должна содержать парфюм-слово И НЕ содержать косметические слова
            cosmetics_match = or_(*cosmetics_exclusions)
            result = query.filter(and_(parfum_match, not_(cosmetics_match)))
            price_logger.info("[PRICE_SEARCH] Applying parfum filter: %d keywords, %d exclusions", 
                            len(PARFUM_KEYWORDS), len(COSMETICS_EXCLUSION_KEYWORDS))
            return result
        else:
            # Если нет исключений, просто ищем парфюм-слова
            price_logger.info("[PRICE_SEARCH] Applying parfum filter with %d keywords", len(PARFUM_KEYWORDS))
            return query.filter(parfum_match)
    elif section == "cosmetics":
        # Косметика: используем существующую логику
        return query.filter(PriceProduct.raw_name.ilike("%Косметика%"))
    elif section == "home":
        return query.filter(PriceProduct.raw_name.ilike("%Для дома%"))
    elif section == "auto":
        return query.filter(PriceProduct.raw_name.ilike("%Автопарфюм%"))
    elif section == "atomizers":
        return query.filter(PriceProduct.raw_name.ilike("%Атомайзер%"))
    elif section == "accessories":
        return query.filter(PriceProduct.raw_name.ilike("%Аксессуар%"))
    return query


def _apply_parfum_filters(query, pf_list: List[str]):
    """
    Применяет внутренние фильтры парфюма (pf) - работает как OR между выбранными
    
    Фильтры:
    - tester: содержит "тестер"
    - sets: содержит "набор", "set", "gift set"
    - analog: содержит "аналог"
    - decant: содержит "отливант", "распив"
    - mini: содержит "миниатюр" ИЛИ (объем <= 10 мл И не отливант)
    """
    if not pf_list:
        return query
    
    from sqlalchemy import or_, and_, not_
    conditions = []
    
    for pf_key in pf_list:
        if pf_key == "mini":
            # Миниатюра: содержит "миниатюр" ИЛИ (объем <= 10 мл И не отливант)
            mini_conditions = []
            
            # Условие 1: содержит "миниатюр"
            mini_keywords = PARFUM_FILTERS.get("mini", [])
            if mini_keywords:
                mini_text_condition = _build_ilike_any(PriceProduct.raw_name, mini_keywords)
                if mini_text_condition:
                    mini_conditions.append(mini_text_condition)
            
            # Условие 2: объем <= 10 мл И не отливант
            # Проверяем volume_value <= 10 и volume_unit в ['мл', 'ml', 'ML']
            volume_condition = and_(
                PriceProduct.volume_value.isnot(None),
                PriceProduct.volume_value <= 10,
                PriceProduct.volume_unit.in_(['мл', 'ml', 'ML', 'Мл', 'Ml'])
            )
            
            # Исключаем отливанты
            decant_keywords = PARFUM_FILTERS.get("decant", [])
            if decant_keywords:
                decant_condition = _build_ilike_any(PriceProduct.raw_name, decant_keywords)
                if decant_condition:
                    volume_condition = and_(volume_condition, not_(decant_condition))
            
            mini_conditions.append(volume_condition)
            
            # Объединяем условия для миниатюры через OR
            if mini_conditions:
                if len(mini_conditions) == 1:
                    conditions.append(mini_conditions[0])
                else:
                    conditions.append(or_(*mini_conditions))
        elif pf_key in PARFUM_FILTERS:
            # Обычные фильтры по ключевым словам
            keywords = PARFUM_FILTERS[pf_key]
            pf_filter = _build_ilike_any(PriceProduct.raw_name, keywords)
            if pf_filter:
                conditions.append(pf_filter)
    
    if not conditions:
        return query
    
    # OR между всеми выбранными фильтрами
    if len(conditions) == 1:
        return query.filter(conditions[0])
    return query.filter(or_(*conditions))


def _search_products(
    db: Session,
    q: str,
    page: int,
    page_size: int,
    upload_id: int = None,
    brand: str | None = None,
    gender: str | None = None,
    ptype: str | None = None,
    psub: str | None = None,
    filter_text: str | None = None,
    section: str | None = None,
    pf: str | None = None,
    hide_decant: bool = False,
):
    price_logger.info("[PRICE_SEARCH] _search_products called with hide_decant=%s (type: %s, value: %s)", hide_decant, type(hide_decant).__name__, repr(hide_decant))
    dialect = db.bind.dialect.name if db.bind else None
    tokens = [t.strip() for t in (q or "").replace(",", " ").split() if t.strip()]
    
    # НЕ устанавливаем section по умолчанию - показываем ВСЕ товары, если section не указан
    # Фильтр применяется только если section явно указан в URL
    if section:
        if isinstance(section, str):
            section = section.strip()
        else:
            section = str(section)
        price_logger.info("[PRICE_SEARCH] Section parameter: '%s'", section)
    else:
        price_logger.info("[PRICE_SEARCH] Section is empty - showing ALL products (no filter)")

    # Попытка FTS5 для SQLite
    fts_ids = []  # Инициализируем переменную
    if dialect == "sqlite" and tokens:
        try:
            # Формируем поисковый запрос для FTS5
            # FTS5 поддерживает операторы: AND, OR, NOT, фразы в кавычках
            # Для простоты используем AND между токенами
            match_expr = " AND ".join(tokens)
            
            # Проверяем существование таблицы FTS5
            fts_exists = db.execute(sa.text("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='price_products_fts5'
            """)).first()
            
            if fts_exists:
                try:
                    # Используем FTS5 через подзапрос
                    fts_subquery = sa.text("""
                        SELECT rowid FROM price_products_fts5 
                        WHERE price_products_fts5 MATCH :match
                    """)
                    
                    # Получаем список ID из FTS5
                    fts_result = db.execute(fts_subquery, {"match": match_expr})
                    fts_ids = [row[0] for row in fts_result] if fts_result else []
                except (OperationalError, sqlite3.DatabaseError) as fts_error:
                    # Если FTS таблица повреждена, логируем и переходим к обычному поиску
                    error_str = str(fts_error)
                    if "malformed" in error_str.lower() or "vtable" in error_str.lower() or "fts" in error_str.lower():
                        price_logger.warning("[PRICE_SEARCH] FTS table is corrupted, falling back to ILIKE search: %s", error_str[:200])
                        fts_exists = False  # Принудительно отключаем FTS
                        fts_ids = []  # Очищаем список
                    else:
                        price_logger.warning("[PRICE_SEARCH] FTS error, falling back to ILIKE: %s", error_str[:200])
                        fts_ids = []  # Очищаем список при любой ошибке
                
                if fts_ids:
                    try:
                        base = (
                            db.query(PriceProduct)
                            .filter(PriceProduct.id.in_(fts_ids))
                        ).filter(
                            PriceProduct.is_in_stock.is_(True),
                            PriceProduct.is_in_current_pricelist.is_(True),
                            PriceProduct.is_active.is_(True),
                        )
                        # Фильтр по бренду для FTS
                        if brand and brand.strip():
                            base = base.filter(PriceProduct.brand.ilike(f"%{brand.strip()}%"))
                        # Фильтр по полу для FTS
                        if gender and gender.strip():
                            gender_value = gender.strip().upper()
                            if gender_value in ['F', 'M', 'U']:
                                base = base.filter(PriceProduct.gender == gender_value)
                        # Фильтр по типу товара (для всех фильтров используем текстовый поиск в raw_name)
                        if ptype and ptype.strip():
                            from app.config_filters import FILTER_TEXT_MAP, PRICE_FILTERS
                            filter_text_value = FILTER_TEXT_MAP.get(ptype.strip())
                            # Если фильтра нет в маппинге, пытаемся найти его label в PRICE_FILTERS
                            if not filter_text_value:
                                for filter_item in PRICE_FILTERS:
                                    if filter_item.get("key") == ptype.strip():
                                        filter_text_value = filter_item.get("label")
                                        break
                            if filter_text_value:
                                # Ищем текст в raw_name (нечувствительно к регистру через ilike)
                                base = base.filter(PriceProduct.raw_name.ilike(f"%{filter_text_value}%"))
                            else:
                                # Если фильтра нет ни в маппинге, ни в PRICE_FILTERS, используем product_type
                                base = base.filter(PriceProduct.product_type == ptype.strip())
                        # Фильтр по подтипу товара
                        if psub and psub.strip():
                            base = base.filter(PriceProduct.product_subtype == psub.strip())
                        # Дополнительный фильтр по тексту (если передан напрямую)
                        if filter_text and filter_text.strip():
                            base = base.filter(PriceProduct.raw_name.ilike(f"%{filter_text.strip()}%"))
                        # Фильтр по разделу (section) - применяется в FTS5 пути только если section указан
                        if section:
                            section_clean = section.strip() if isinstance(section, str) else str(section)
                            price_logger.info("[PRICE_SEARCH] [FTS5] Applying section filter: '%s'", section_clean)
                            section_filter = _apply_section_filter(base, section_clean)
                            if section_filter is not None:
                                base = section_filter
                                count_after_section = base.count()
                                price_logger.info("[PRICE_SEARCH] [FTS5] Products after section filter: %s", count_after_section)
                            else:
                                price_logger.warning("[PRICE_SEARCH] [FTS5] Section filter returned None")
                        else:
                            price_logger.error("[PRICE_SEARCH] [FTS5] Section filter NOT applied! section='%s' (type: %s)", section, type(section))
                        # Фильтр по внутренним фильтрам парфюма (pf)
                        if section and section.strip() == "parfum" and pf:
                            pf_list = _parse_csv_param(pf)
                            if pf_list:
                                pf_filter = _apply_parfum_filters(base, pf_list)
                                if pf_filter is not None:
                                    base = pf_filter
                        # Фильтр для скрытия товаров с "отливант" в названии (нечувствительно к регистру)
                        # Применяем ПОСЛЕ всех поисковых условий в FTS5 пути, чтобы исключить найденные товары
                        # Ищем в raw_name и product_name, так как "отливант" может быть в разных полях
                        if hide_decant:
                            from sqlalchemy import not_, or_
                            price_logger.info("[PRICE_SEARCH] [FTS5] Applying hide_decant filter AFTER all filters: excluding products with 'отливант' in raw_name or product_name")
                            count_before_filter_fts = base.count()
                            price_logger.info("[PRICE_SEARCH] [FTS5] Products count BEFORE hide_decant filter: %s", count_before_filter_fts)
                            # Используем ilike с обоими вариантами регистра, так как ilike в SQLite может быть чувствителен к регистру кириллицы
                            base = base.filter(
                                not_(
                                    or_(
                                        PriceProduct.raw_name.ilike("%отливант%"),
                                        PriceProduct.raw_name.ilike("%ОТЛИВАНТ%"),
                                        PriceProduct.product_name.ilike("%отливант%"),
                                        PriceProduct.product_name.ilike("%ОТЛИВАНТ%"),
                                    )
                                )
                            )
                            count_after_filter_fts = base.count()
                            price_logger.info("[PRICE_SEARCH] [FTS5] Products count AFTER hide_decant filter: %s (excluded: %s)", count_after_filter_fts, count_before_filter_fts - count_after_filter_fts)
                        if upload_id:
                            base = base.filter(
                                sa.exists().where(
                                    (PriceHistory.price_upload_id == upload_id)
                                    & (PriceHistory.price_product_id == PriceProduct.id)
                                )
                            )
                        total = base.count()
                        items = (
                            base.order_by(PriceProduct.id.desc())
                            .offset((page - 1) * page_size)
                            .limit(page_size)
                            .all()
                        )
                        return items, total
                    except Exception as fts_query_error:
                        price_logger.warning("[PRICE_SEARCH] FTS query failed, falling back to ILIKE: %s", fts_query_error)
                        # Продолжаем к fallback поиску
                else:
                    # FTS не вернул результатов, переходим к обычному поиску
                    price_logger.debug("[PRICE_SEARCH] FTS returned no results, using ILIKE search")
        except Exception as e:
            price_logger.exception("FTS5 search failed, fallback to ILIKE: %s", e)

    # Fallback: обычные ILIKE
    base_query = db.query(PriceProduct).filter(
        PriceProduct.is_in_stock.is_(True),
        PriceProduct.is_in_current_pricelist.is_(True),
        PriceProduct.is_active.is_(True),
    )
    if upload_id:
        base_query = base_query.filter(
            sa.exists().where(
                (PriceHistory.price_upload_id == upload_id)
                & (PriceHistory.price_product_id == PriceProduct.id)
            )
        )
    # Фильтр по бренду
    if brand and brand.strip():
        base_query = base_query.filter(PriceProduct.brand.ilike(f"%{brand.strip()}%"))
    # Фильтр по полу
    if gender and gender.strip():
        gender_value = gender.strip().upper()
        if gender_value in ['F', 'M', 'U']:
            base_query = base_query.filter(PriceProduct.gender == gender_value)
    # Фильтр по типу товара (для всех фильтров используем текстовый поиск в raw_name)
    if ptype and ptype.strip():
        from app.config_filters import FILTER_TEXT_MAP, PRICE_FILTERS
        filter_text_value = FILTER_TEXT_MAP.get(ptype.strip())
        # Если фильтра нет в маппинге, пытаемся найти его label в PRICE_FILTERS
        if not filter_text_value:
            for filter_item in PRICE_FILTERS:
                if filter_item.get("key") == ptype.strip():
                    filter_text_value = filter_item.get("label")
                    break
        if filter_text_value:
            # Ищем текст в raw_name (нечувствительно к регистру через ilike)
            price_logger.info("[PRICE_SEARCH] Applying filter ptype=%s, searching for text '%s' in raw_name", ptype.strip(), filter_text_value)
            base_query = base_query.filter(PriceProduct.raw_name.ilike(f"%{filter_text_value}%"))
        else:
            # Если фильтра нет ни в маппинге, ни в PRICE_FILTERS, используем product_type
            price_logger.info("[PRICE_SEARCH] Applying filter ptype=%s, using product_type field", ptype.strip())
            base_query = base_query.filter(PriceProduct.product_type == ptype.strip())
    # Фильтр по подтипу товара
    if psub and psub.strip():
        base_query = base_query.filter(PriceProduct.product_subtype == psub.strip())
    # Дополнительный фильтр по тексту (если передан напрямую)
    if filter_text and filter_text.strip():
        base_query = base_query.filter(PriceProduct.raw_name.ilike(f"%{filter_text.strip()}%"))
    # Фильтр по разделу (section) - применяется только если section явно указан
    if section:
        section_clean = section.strip() if isinstance(section, str) else str(section)
        price_logger.info("[PRICE_SEARCH] Applying section filter: '%s' (original: '%s')", section_clean, section)
        section_filter = _apply_section_filter(base_query, section_clean)
        if section_filter is not None:
            base_query = section_filter
            # УБРАЛИ count() - он слишком медленный, логируем только факт применения фильтра
            price_logger.info("[PRICE_SEARCH] Section filter applied: '%s'", section_clean)
        else:
            price_logger.warning("[PRICE_SEARCH] Section filter returned None for section: '%s'", section_clean)
    else:
        price_logger.error("[PRICE_SEARCH] Section filter NOT applied! section='%s' (type: %s)", section, type(section))
    # Фильтр по внутренним фильтрам парфюма (pf)
    if section and section.strip() == "parfum" and pf:
        pf_list = _parse_csv_param(pf)
        if pf_list:
            price_logger.info("[PRICE_SEARCH] Applying parfum filters: %s", pf_list)
            pf_filter = _apply_parfum_filters(base_query, pf_list)
            if pf_filter is not None:
                base_query = pf_filter
                # УБРАЛИ count() - он слишком медленный
                price_logger.info("[PRICE_SEARCH] Parfum filters applied: %s", pf_list)
    # Используем search_text если есть, иначе обычные поля
    # Важно: если q пустой, но есть ptype, мы все равно должны показать результаты
    if tokens:
        if hasattr(PriceProduct, 'search_text') and PriceProduct.search_text:
            # Используем search_text для поиска
            for tok in tokens:
                like = f"%{tok}%"
                base_query = base_query.filter(
                    (PriceProduct.search_text.ilike(like))
                    | (PriceProduct.external_article.ilike(like))
                    | (PriceProduct.raw_name.ilike(like))
                    | (PriceProduct.product_name.ilike(like))
                )
        else:
            # Fallback на обычные поля
            for tok in tokens:
                like = f"%{tok}%"
                base_query = base_query.filter(
                    (PriceProduct.external_article.ilike(like))
                    | (PriceProduct.raw_name.ilike(like))
                    | (PriceProduct.product_name.ilike(like))
                )
    # Фильтр для скрытия товаров с "отливант" в названии (нечувствительно к регистру)
    # Применяем ПОСЛЕ поиска по токенам, чтобы исключить найденные товары с "отливант"
    # Ищем в raw_name и product_name, так как "отливант" может быть в разных полях
    if hide_decant:
        from sqlalchemy import not_, func, or_
        price_logger.info("[PRICE_SEARCH] Applying hide_decant filter: excluding products with 'отливант'")
        # УБРАЛИ count() до фильтра - слишком медленно
        # Используем ilike с обоими вариантами регистра, так как ilike в SQLite может быть чувствителен к регистру кириллицы
        base_query = base_query.filter(
            not_(
                or_(
                    PriceProduct.raw_name.ilike("%отливант%"),
                    PriceProduct.raw_name.ilike("%ОТЛИВАНТ%"),
                    PriceProduct.product_name.ilike("%отливант%"),
                    PriceProduct.product_name.ilike("%ОТЛИВАНТ%"),
                )
            )
        )
        # УБРАЛИ count() после фильтра и проверку sample_items - слишком медленно
        price_logger.info("[PRICE_SEARCH] hide_decant filter applied")
    else:
        price_logger.info("[PRICE_SEARCH] hide_decant is False, filter NOT applied")
    # Если q пустой, но есть ptype, показываем все товары, соответствующие фильтру
    price_logger.info("[PRICE_SEARCH] Query: q='%s', ptype='%s', tokens=%s, hide_decant=%s, final query will filter by ptype if set", q, ptype, len(tokens), hide_decant)
    
    # ОПТИМИЗАЦИЯ: Для больших результатов используем приблизительный подсчет
    # Если нет фильтров (пустой запрос), ограничиваем count для производительности
    MAX_COUNT_FOR_EXACT = 50000  # Если больше - используем приблизительный подсчет
    
    # Сначала получаем товары (это быстрее чем count для больших таблиц)
    items = (
        base_query.order_by(PriceProduct.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    
    # Если есть фильтры (q, brand, ptype и т.д.), делаем точный count
    # Если фильтров нет и результатов много, используем приблизительный подсчет
    has_filters = bool(tokens or brand or gender or ptype or psub or section or pf or hide_decant or upload_id)
    
    if has_filters:
        # Есть фильтры - делаем точный подсчет (но с таймаутом)
        try:
            total = base_query.count()
        except Exception as count_error:
            price_logger.warning("[PRICE_SEARCH] Count failed, using estimate: %s", count_error)
            # Приблизительный подсчет через LIMIT
            total = min(base_query.limit(MAX_COUNT_FOR_EXACT + 1).count(), MAX_COUNT_FOR_EXACT)
            if total == MAX_COUNT_FOR_EXACT:
                total = MAX_COUNT_FOR_EXACT  # Помечаем как "больше чем"
    else:
        # Нет фильтров - используем приблизительный подсчет для производительности
        # Проверяем, есть ли хотя бы один результат
        test_count = base_query.limit(1).count()
        if test_count == 0:
            total = 0
        else:
            # Используем приблизительный подсчет через LIMIT
            total = min(base_query.limit(MAX_COUNT_FOR_EXACT + 1).count(), MAX_COUNT_FOR_EXACT)
            if total == MAX_COUNT_FOR_EXACT:
                total = MAX_COUNT_FOR_EXACT  # Показываем "больше чем MAX_COUNT"
    
    price_logger.info("[PRICE_SEARCH] Total products found: %s (has_filters=%s)", total, has_filters)
    return items, total


def _last_history_map(db: Session, product_ids: List[int]) -> Dict[int, PriceHistory]:
    if not product_ids:
        return {}
    sub = (
        db.query(
            PriceHistory.price_product_id.label("pid"),
            sa.func.max(PriceHistory.created_at).label("max_created"),
        )
        .filter(PriceHistory.price_product_id.in_(product_ids))
        .group_by(PriceHistory.price_product_id)
        .subquery()
    )
    rows = (
        db.query(PriceHistory)
        .join(sub, (PriceHistory.price_product_id == sub.c.pid) & (PriceHistory.created_at == sub.c.max_created))
        .all()
    )
    return {r.price_product_id: r for r in rows}


def _latest_price_map(db: Session, product_ids: List[int], field: str = "price_2") -> Dict[int, Decimal]:
    if not product_ids:
        return {}
    sub = (
        db.query(
            PriceHistory.price_product_id.label("pid"),
            sa.func.max(PriceHistory.created_at).label("max_created"),
        )
        .filter(PriceHistory.price_product_id.in_(product_ids))
        .group_by(PriceHistory.price_product_id)
        .subquery()
    )
    rows = (
        db.query(
            PriceHistory.price_product_id,
            PriceHistory.new_price_2,
            PriceHistory.new_price_1,
            PriceHistory.price,
        )
        .join(sub, (PriceHistory.price_product_id == sub.c.pid) & (PriceHistory.created_at == sub.c.max_created))
        .all()
    )
    result = {}
    for pid, new_price_2, new_price_1, price in rows:
        if field == "price_1":
            val = new_price_1 if new_price_1 is not None else new_price_2 if new_price_2 is not None else price
        else:
            val = new_price_2 if new_price_2 is not None else price
        result[pid] = Decimal(val or 0)
    return result


def _to_float(val):
    try:
        if val is None or val == "":
            return None
        return float(val)
    except Exception:
        return None


# Роуты ревью нормализации перенесены в app/routes/normalization.py
# GET /normalization/review
# POST /normalization/review/{product_id}
# POST /normalization/search-fragella/{product_id}


def _extract_brand_alias_from_raw(raw_name: str, model_name: Optional[str] = None) -> str:
    """
    Извлекает кандидата алиаса бренда из raw_name.
    Берет часть строки от начала до начала модели или до стоп-слов.
    """
    if not raw_name:
        return ""
    
    import re
    raw_name = raw_name.strip()
    
    # Стоп-слова для определения конца названия бренда
    stop_words = {
        'унисекс', 'женск', 'женский', 'мужск', 'мужской',
        'парф', 'парфюмерная', 'туалет', 'туалетная', 'вода',
        'мл', 'ml', 'г', 'гр', 'g', 'gr',
        'тестер', '(тестер)', 'отливант', 'пробник', 'sample',
        'духи', 'edp', 'edt', 'eau', 'de', 'parfum', 'toilette',
        'миниатюра', 'mini', 'decant', 'for', 'women', 'men'
    }
    
    # Если есть > → всё до первого >
    if '>' in raw_name:
        parts = raw_name.split('>', 1)
        candidate = parts[0].strip()
        if candidate:
            return candidate
    
    # Если есть model_name, пытаемся найти его в raw_name и взять всё до него
    if model_name and model_name.strip():
        model_lower = model_name.strip().lower()
        raw_lower = raw_name.lower()
        
        # Ищем модель в raw_name (может быть с разным регистром)
        model_pos = raw_lower.find(model_lower)
        if model_pos > 0:
            # Берем часть до модели
            candidate = raw_name[:model_pos].strip()
            # Очищаем от лишних символов
            candidate = re.sub(r'[^\w&\-\.\s]+$', '', candidate).strip()
            if candidate and len(candidate) >= 2:
                return candidate
    
    # Иначе берем первые 2-5 слов до стоп-слов
    words = raw_name.split()
    
    # Ищем позицию первого стоп-слова
    stop_pos = None
    for i, word in enumerate(words):
        word_lower = word.lower().strip('.,;:()[]{}')
        if word_lower in stop_words:
            stop_pos = i
            break
    
    # Берем от 2 до 5 слов до стоп-слова (или все слова, если стоп-слов нет)
    if stop_pos is not None:
        max_words = min(5, stop_pos)
    else:
        max_words = min(5, len(words))
    
    # Минимум 2 слова для многословных брендов
    if max_words < 2:
        max_words = min(2, len(words))
    
    if max_words == 0:
        return ""
    
    candidate_words = words[:max_words]
    candidate = ' '.join(candidate_words).strip()
    
    # Очищаем от лишних символов в начале/конце
    candidate = re.sub(r'^[^\w&\-\.]+|[^\w&\-\.]+$', '', candidate)
    
    return candidate if candidate and len(candidate) >= 2 else ""


# Роут сохранения ревью нормализации перенесён в app/routes/normalization.py


@router.post("/normalize/batch", response_class=JSONResponse)
async def normalize_batch(
    request: Request,
    upload_id: Optional[int] = Form(None),
    limit: int = Form(100),
    current_user: User = Depends(require_roles(["ADMIN"])),
    db: Session = Depends(get_db),
):
    """Запуск нормализации для товаров со статусом pending"""
    try:
        # Получаем товары для нормализации
        query = db.query(PriceProduct).filter(PriceProduct.ai_status == "pending")
        
        if upload_id:
            # Нормализуем только товары из конкретной загрузки
            # Используем JOIN вместо IN для избежания ошибки "too many SQL variables"
            query = query.join(PriceHistory, PriceProduct.id == PriceHistory.price_product_id).filter(
                PriceHistory.price_upload_id == upload_id
            ).distinct()
        
        products = query.limit(limit).all()
        
        if not products:
            return {
                "success": True,
                "processed": 0,
                "message": "Нет товаров для нормализации"
            }
        
        processed = 0
        errors = 0
        
        for product in products:
            try:
                if not product.raw_name:
                    product.ai_status = "error"
                    product.normalization_notes = "Нет raw_name"
                    continue
                
                # Выполняем нормализацию
                normalized = normalize_price_row(product.raw_name, db)
                
                if normalized:
                    product.norm_brand = normalized.brand
                    product.brand_confidence = normalized.brand_confidence
                    product.model_name = normalized.model_name
                    product.series = normalized.series
                    product.category_path_json = json.dumps(normalized.category_path, ensure_ascii=False) if normalized.category_path else None
                    product.attrs_json = json.dumps(normalized.attrs, ensure_ascii=False) if normalized.attrs else None
                    product.ai_group_key = normalized.group_key
                    product.variant_key = normalized.variant_key
                    product.search_text = normalized.search_text
                    product.normalization_notes = normalized.notes
                    product.ai_status = "review" if normalized.needs_review else "ok"
                    # Сохраняем product_type и product_subtype из attrs
                    if normalized.attrs:
                        product.product_type = normalized.attrs.get('product_type')
                        product.product_subtype = normalized.attrs.get('product_subtype')
                    
                    # Обновляем каталог
                    if normalized.group_key:
                        try:
                            upsert_catalog_from_price(product, normalized, db)
                        except Exception as e:
                            price_logger.warning("Catalog upsert failed for product_id=%s: %s", product.id, str(e)[:200])
                    
                    processed += 1
                else:
                    product.ai_status = "error"
                    product.normalization_notes = "Ошибка нормализации"
                    errors += 1
            except Exception as e:
                price_logger.exception("Normalization failed for product_id=%s: %s", product.id, e)
                product.ai_status = "error"
                product.normalization_notes = f"Ошибка: {str(e)[:200]}"
                errors += 1
        
        db.commit()
        
        return {
            "success": True,
            "processed": processed,
            "errors": errors,
            "total": len(products),
            "message": f"Обработано: {processed}, ошибок: {errors}"
        }
    except Exception as e:
        price_logger.exception("Batch normalization failed: %s", e)
        db.rollback()
        return {
            "success": False,
            "error": str(e)[:200]
        }


@router.post("/{product_id}/to_catalog", response_class=JSONResponse)
async def add_to_catalog(
    product_id: int,
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    """Добавляет товар из прайса в каталог"""
    try:
        product = db.query(PriceProduct).filter(PriceProduct.id == product_id).first()
        if not product:
            return {"success": False, "error": "Товар не найден"}
        
        if not product.raw_name:
            return {"success": False, "error": "У товара нет raw_name"}
        
        # Проверяем, есть ли уже нормализация
        if not product.norm_brand or not product.model_name:
            # Пытаемся нормализовать
            normalized = normalize_price_row(product.raw_name, db)
            if normalized:
                product.norm_brand = normalized.brand
                product.brand_confidence = normalized.brand_confidence
                product.model_name = normalized.model_name
                product.series = normalized.series
                product.category_path_json = json.dumps(normalized.category_path, ensure_ascii=False) if normalized.category_path else None
                product.attrs_json = json.dumps(normalized.attrs, ensure_ascii=False) if normalized.attrs else None
                product.ai_group_key = normalized.group_key
                product.variant_key = normalized.variant_key
                product.search_text = normalized.search_text
                product.normalization_notes = normalized.notes
                product.ai_status = "review" if normalized.needs_review else "ok"
                # Сохраняем product_type и product_subtype из attrs
                if normalized.attrs:
                    product.product_type = normalized.attrs.get('product_type')
                    product.product_subtype = normalized.attrs.get('product_subtype')
                db.commit()
            else:
                return {"success": False, "error": "Не удалось нормализовать товар. Запустите нормализацию сначала."}
        
        # Создаем NormalizedResult из полей продукта
        normalized = NormalizedResult(
            brand=product.norm_brand,
            brand_confidence=float(product.brand_confidence or 0),
            model_name=product.model_name or "",
            series=product.series,
            category_path=json.loads(product.category_path_json) if product.category_path_json else None,
            attrs=json.loads(product.attrs_json) if product.attrs_json else None,
            group_key=product.ai_group_key,
            variant_key=product.variant_key,
            search_text=product.search_text,
            notes=product.normalization_notes,
            needs_review=(product.ai_status == "review")
        )
        
        # Добавляем в каталог
        try:
            catalog_variant = upsert_catalog_from_price(product, normalized, db)
            db.commit()
            
            catalog_item_id = None
            if catalog_variant:
                catalog_item_id = catalog_variant.catalog_item_id
            
            return {
                "success": True,
                "catalog_item_id": catalog_item_id,
                "message": "Товар добавлен в каталог"
            }
        except IntegrityError as ie:
            db.rollback()
            price_logger.warning("Catalog upsert IntegrityError for product_id=%s: %s", product_id, str(ie)[:200])
            return {"success": False, "error": "Ошибка при добавлении в каталог (возможно, уже существует)"}
        except Exception as e:
            db.rollback()
            price_logger.exception("Catalog upsert failed for product_id=%s: %s", product_id, e)
            return {"success": False, "error": f"Ошибка: {str(e)[:200]}"}
    except Exception as e:
        price_logger.exception("Add to catalog failed for product_id=%s: %s", product_id, e)
        db.rollback()
        return {"success": False, "error": str(e)[:200]}


@router.get("/product/{product_id}", response_class=HTMLResponse)
async def price_product_detail(
    request: Request,
    product_id: int,
    current_user: User = Depends(require_permission("price.search")),
    db: Session = Depends(get_db),
):
    product = db.query(PriceProduct).filter(PriceProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Товар не найден")
    history = (
        db.query(PriceHistory)
        .filter(PriceHistory.price_product_id == product.id)
        .order_by(PriceHistory.created_at.desc())
        .limit(200)
        .all()
    )
    orders = []  # Заказы пока не используем
    can_view_cost = user_has_permission(current_user, db, "prices.view_cost")
    can_view_client = user_has_permission(current_user, db, "prices.view_client") or can_view_cost
    can_view_margin = user_has_permission(current_user, db, "prices.view_margin")

    partner_pricing = None
    history_with_partner_prices = None
    if getattr(current_user, "partner_id", None) and can_view_client and history:
        last = history[0]
        raw_base_price = Decimal((last.new_price_1 if last.new_price_1 is not None else last.new_price_2 if last.new_price_2 is not None else last.price) or 0)
        partner_id = current_user.partner_id
        policy = get_partner_pricing_policy(db, partner_id)
        admin_pct = policy.admin_markup_percent
        total_pct = get_total_markup_percent(db, partner_id)
        base_price = calc_partner_price(raw_base_price, policy.partner_price_markup_percent)
        partner_pricing = {
            "admin_markup_percent": admin_pct,
            "partner_price_markup_percent": policy.partner_price_markup_percent,
            "partner_markup_percent": (total_pct - admin_pct) if total_pct is not None else None,
            "total_markup_percent": total_pct,
            "base_price": base_price,
            "admin_price": calc_client_price(base_price, admin_pct),
            "total_price": calc_client_price(base_price, total_pct),
        }
        
        # Вычисляем цены партнера для каждой записи истории
        history_with_partner_prices = []
        for h in history:
            # Берем базовую цену из истории (price_1 или price_2 или price)
            raw_price = Decimal((h.new_price_1 if h.new_price_1 is not None else h.new_price_2 if h.new_price_2 is not None else h.price) or 0)
            # Вычисляем цену партнера с учетом надбавки
            partner_price = calc_partner_price(raw_price, policy.partner_price_markup_percent)
            history_with_partner_prices.append({
                "history": h,
                "partner_price": partner_price,
            })
    
    return templates.TemplateResponse(
        "price_product_detail.html",
        {
            "request": request,
            "current_user": current_user,
            "product": product,
            "history": history,
            "history_with_partner_prices": history_with_partner_prices,
            "orders": orders,
            "can_view_cost": can_view_cost,
            "can_view_client": can_view_client,
            "can_view_margin": can_view_margin,
            "partner_pricing": partner_pricing,
        },
    )
